import os
import time
import traceback
import chardet
from datetime import date, datetime
from wsgiref import headers

from openpyxl import Workbook, load_workbook
# 导入字体、颜色、对齐、填充模块、边框、侧边、自动换行
from openpyxl.styles import Font, colors, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image    # 导入图片模块
from pywebio import start_server
from pywebio.input import *
from pywebio.output import *
from pywebio.pin import *
from pywebio.session import *

import 获取系统信息 as GD  # 自己写的其它模块

'''
内容未改动，除了公司的系统内容以外，所有代码都在这里了，不涉及我们系统的代码都可以正常运行。酱友们可以自行参考，也可以留言或交流群里讨论，互相学习，一起成长。
'''

# 检测文件编码
def txt_x(file_path):
    with open(file_path, "rb") as f:
        msg = f.read()
        # 光标返回到开始
        f.seek(0)
        result = chardet.detect(msg)
        # print(result['encoding'])
        return result['encoding']

# 合并TXT文本
def 文本批量合并(files):
    txts = []
    h = 1  # 标记文件数，从第2个开始不取标题
    put_processbar('file', auto_close=True)
    for file in files:  #遍历文件夹
        set_processbar('file', h / len(files), label=h)
        # 判断是不是txt文件
        if file[-4:] in ['.txt', '.TXT']:
            # print(f'正在加载{file}文件')
            try:
                # 打开文件
                with open(file, encoding='GB18030') as f:
                    # 读取每行并制作成列表赋值给lines
                    lines = f.readlines()

            except:
                try:
                    coding = txt_x(file)
                    with open(file, encoding=coding) as f:
                        lines = f.readlines()
                except Exception as e:
                    # 输出错误提示
                    print(datetime.now())
                    print(traceback.format_exc())
                    print('====='*50)
                    print(e)
                    lines = ''
                    put_text('文本编码未识别')

        else:
            put_text('你选择的不是文本文件')
            continue
        if not lines:
            put_text('没有内容')
            continue

        num = 1
        for line in lines:
            if h == 1:
                txts.append(line)
                continue
            if num == 1:
                num += 1
                continue
            txts.append(line)
        h += 1
        # put_text('文本加载完成')
    return txts

# 转换成表格
def txt_xlsx(txts, name):
    # 创建表格
    wb = Workbook()
    ws = wb.active
    x = 1   # 记录行数
    year_19 = 0   # 记录年份条数
    year_20 = 0   # 记录年份条数
    year_21 = 0   # 记录年份条数
    year_22 = 0   # 记录年份条数
    j = 0   # 记录拒付行数
    put_processbar('line', auto_close=True)
    for line in txts:
        set_processbar('line', x / len(txts), label=x)
        lst = []
        if not line:
            break
        if len(line.split('|')) == 1:
            切割符 = '\t'
        else:
            切割符 = '|'

        try:
            if line.split(切割符)[5][:4] < '2020':
                year_19 += 1
            elif line.split(切割符)[5][:4] == '2020':
                year_20 += 1
            elif line.split(切割符)[5][:4] == '2021':
                year_21 += 1
            elif line.split(切割符)[5][:4] == '2022':
                year_22 += 1
        except:
            j += 1

        for i in range(len(line.split(切割符))):
            item = line.split(切割符)
            if len(item) < 20:  # 判断是不是拒付内容
                j += 1
                break
            item = item[i].strip()  # 去除空格和回车
            if x == 0:
                lst.append(item)
            else:
                if i > 2:  # 检测身份证号(每行第2列)
                    if is_number(item):
                        item = float(item)
                        lst.append(item)
                    else:
                        lst.append(item)

                else:
                    lst.append(item)
        if lst:
            ws.append(lst)
        x += 1
    put_text(f'一共处理了{x-2}行数据')
    put_text(f'其中{year_19}条小于20年，{year_20}条20年，{year_21}条21年，{year_22}条22年，{j}条拒付内容')
    wb.save(f'缓存文件夹/{name}.xlsx')

## 判断是否是数字
def is_number(s):
    try:  # 如果能运行float(s)语句，返回True（字符串s是浮点数）
        float(s)
        return True
    except ValueError:  # ValueError为Python的一种标准异常，表示"传入无效的参数"
        return False

# 基础信息查询函数
def 导入批量信息(loadfiles):
    ls = []
    row_nmb = 0 # 表格行数
    空行 = 0
    wb = load_workbook(loadfiles)
    ws = wb.active   # 获取活跃sheet表
    put_processbar('name', auto_close=True)
    for row in ws.rows:
        row_nmb += 1
        set_processbar('name', row_nmb / ws.max_row, label=row_nmb)
        # 检查空行
        c = 0
        for cell in row:
            if cell.value is not None:
                c = 1
        if c == 0:
            空行 += 1
            continue
        
        if row_nmb == 1:
            continue

        ls.append(row[0].value)
                
    put_text(f'共导入{row_nmb-空行-1}条数据。')
    # print(ls)
    return ls

def 导入团批信息(loadfiles):
    # put_text('正在导入团批信息。。。')
    di = {}  # 保存团批内容
    row_nmb = 0 # 表格行数
    空行 = 0
    wb = load_workbook(loadfiles)
    ws = wb.active   # 获取活跃sheet表
    put_processbar('团批', auto_close=True)
    for row in ws.rows:
        row_nmb += 1
        set_processbar('团批', row_nmb / ws.max_row, label=row_nmb)
        # 检查空行
        c = 0
        if row_nmb == 1:
            continue

        for cell in row:
            if cell.value is not None:
                c = 1
        if c == 0:
            空行 += 1
            continue
        
        # 0序号，1姓名，2身份证号，3案件号，4票据数]
        # 序号:[0姓名，1身份证号，2案件号，3票据数]
        di[row[0].value] = [row[1].value, row[2].value, row[4].value]
                
    put_text(f'共导入{ws.max_row-1}条数据')
    return di

def 检查基础字段(批次号, 团批, 单位):
    wb = Workbook()     # 创建新工作薄
    ws = wb.active      # 获取活跃sheet表
    # 红色 = ['#ff0000'] # 设置红色
    fille = PatternFill('solid', fgColor='ffc7ce')
    # font = Font(u'微软雅黑', size=11, bold=False, italic=False, strike=False, color='ffc7ce')
    font = Font(color='ff0000')
    
    if 团批:
        title = ['序号', '姓名', '案件号', '姓名', '身份证号', '票据数', '电话', '银行卡', '起付线', '超封顶', '自费', '住院', '门特', '错误提示', '退单', '票据年份', '校验保单', '应选保单号', '综合多样化', '问题件', '审核员', '备注']
        默认 = ['-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-']
        ws.append(title)    # 批量添加标题
        减列数 = 0
        ws.cell(row=1, column=14-减列数, value='错误提示').fill = fille
    else:
        title = ['序号', '姓名', '案件号', '电话', '银行卡', '起付线', '超封顶', '自费', '住院', '门特', '错误提示', '退单', '票据年份', '校验保单', '应选保单号', '综合多样化', '问题件', '审核员', '备注']
        默认 = ['-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-', '-']
        ws.append(title)    # 批量添加标题
        减列数 = 3
        ws.cell(row=1, column=14-减列数, value='错误提示').fill = fille
            
    # 获取案件信息
    url = GD.批次号查询网址(批次号)
    data1 = GD.获取案件信息(url, headers)

    if data1 == '没有更多啦~':
        put_text(f'系统没有查询到此批次号{批次号}')
        return data1

    案件总数 = data1['page']['count']
    案件总页数 = data1['page']['pages']
    put_text(f'批次{批次号}共{案件总数}件案件。\n')
    data2 = data1['data']
    if 案件总页数 > 1:
        # 按页循环获取数据
        for page in range(2, 案件总页数+1):
            url = GD.批次号查询网址(批次号, 10, page)
            data1 = GD.获取案件信息(url, headers)   # 获取案件列表页
            # 把每页的数据合并到一起
            for i in data1['data']:
                data2.append(i)

    if data2:
        序号 = 1    # 案件位置
        put_processbar(批次号, auto_close=True)
        for x in data2:
            set_processbar(批次号, 序号 / 案件总数, label=序号)
            time.sleep(0.1)
            序号 += 1
            票据年份 = ''
            错误标记汇总 = ''
            退单原因汇总 = ''
            退单原因 = ''
            票据号 = []
            ws.append(默认)    # 添加默认内容
            个案列表 = GD.提取案件列表个案详情(x)

            if 个案列表['核查校验'] == 0 or 个案列表['案件状态'] == '已分析':
                错误标记汇总 = '未核查;'

            # 获取基础信息
            url = GD.案件详情查询网址(个案列表['案件id'])
            个案信息 = GD.获取案件信息(url, headers)   # 获取个案详情
            基础信息 = GD.提取案件详情基础信息(个案信息)

            if 团批:
                案件位置 = 基础信息['案件号'][len(批次号):]
                if 案件位置 in 团批:
                    团批内容 = 团批[案件位置]
                    
                    # 序号:[0姓名，1身份证号，2票据数]
                    if 基础信息['身份证号'] == 团批内容[1]:    # 检测身份证号
                        ws[f'E{序号}'] = '*'

                        if 基础信息['姓名'] == 团批内容[0]: # 检测姓名
                            ws[f'D{序号}'] = '*'
                        else:
                            ws.cell(row=序号, column=4, value="错").fill = fille
                    else:
                        ws.cell(row=序号, column=5, value="错").fill = fille

                    门诊票数 = 住院票数 = 门特票数 = 0
                    if 个案信息['stub']:
                        门诊票数 = len(个案信息['stub'])
                    if 个案信息['stub_hospital']:
                        住院票数 = len(个案信息['stub_hospital'])
                    if 个案信息['stub_whole']:
                        门特票数 = len(个案信息['stub_whole'])

                    if 门诊票数+住院票数+门特票数 == 团批内容[2]:   # 检测票数
                        ws[f'F{序号}'] = '*'
                    else:
                        ws.cell(row=序号, column=6, value=门诊票数+住院票数+门特票数-int(团批内容[2])).font = font
            # 各列行数
            ws[f'A{序号}'] = 序号 - 1
            ws[f'B{序号}'] = 个案列表['姓名']
            ws[f'C{序号}'] = 个案列表['案件号']
            ws.cell(row=序号, column=21-减列数, value=个案列表['审核员'])
            
            # 查询本单位保单号
            单位信息 = 单位简称获取(批次号)   # 0是单位简称，1是保单号简称，2是保单号全称
            ws.cell(row=序号, column=18-减列数, value=单位信息[1])

            try:
                if 基础信息['联系电话']:
                    if is_number(基础信息['联系电话']):
                        if len(基础信息['联系电话']) != 11:
                            ws.cell(row=序号, column=7-减列数, value="位数不对").fill = fille

                        elif len(基础信息['联系电话']) == 11:
                            ws.cell(row=序号, column=7-减列数, value="*")
                    else:
                        ws.cell(row=序号, column=7-减列数, value="错误").fill = fille
            except:
                ws.cell(row=序号, column=7-减列数, value="错误").fill = fille

            if 基础信息['保单银行账号']:
                if is_number(基础信息['保单银行账号']):
                    ws.cell(row=序号, column=8-减列数, value='*')
                else:
                    ws.cell(row=序号, column=8-减列数, value='无').fill = fille
            else:
                ws.cell(row=序号, column=8-减列数, value='无').fill = fille

            # 判断保单号，21年是6856，22年是0374
            if 基础信息['保单号']:
                已选择保单数量 = 基础信息['保单号'].split(',')
                for i in 已选择保单数量:
                    if i in 单位信息[2]:
                        ws.cell(row=序号, column=17-减列数, value='*')
                    else:
                        ws.cell(row=序号, column=17-减列数, value='保单错误').fill = fille
            else:
                ws.cell(row=序号, column=17-减列数, value='无').fill = fille

            if 基础信息['特殊人员标识']:
                ws.cell(row=序号, column=19-减列数, value=基础信息['特殊人员标识']).fill = fille
            elif '退休' in 基础信息['保单方案'] or '退职' in 基础信息['保单方案']:
                pass
            else:
                ws.cell(row=序号, column=19-减列数, value='在职').fill = fille

            if not 基础信息.get('生效时间'):
                ws.cell(row=序号, column=19-减列数, value='无承保').fill = fille

            if 基础信息['问题件'] == '是':
                ws.cell(row=序号, column=20-减列数, value='是:'+基础信息['问题件简述']).fill = fille
                
            # 是否有增值税
            if 个案信息['stub_invoice']:
                错误标记汇总 = 错误标记汇总+'有增值税栏;'

            # 是否有医保未结算
            if 个案信息['stub_none']:
                错误标记汇总 = 错误标记汇总+'有未医保栏;'

            ws.cell(row=序号, column=22-减列数, value=基础信息['备注'])

            # 检测票据明细内容
            # 是否有门诊
            if 个案信息['stub']:
                门诊合计 = GD.提取案件详情城镇门诊合计信息(个案信息['sum_stub'])

                if '退休' in 基础信息['保单方案'] or '退职' in 基础信息['保单方案']:
                    if 门诊合计['起付金额'] > 1300:
                        ws.cell(row=序号, column=9-减列数, value=门诊合计['起付金额']).font = font
                    else:
                        ws.cell(row=序号, column=9-减列数, value=门诊合计['起付金额'])
                else:
                    if 门诊合计['起付金额'] > 1800:
                        ws.cell(row=序号, column=9-减列数, value=门诊合计['起付金额']).font = font
                    else:
                        ws.cell(row=序号, column=9-减列数, value=门诊合计['起付金额'])

                if 门诊合计['超封顶金额'] > 0:
                    ws.cell(row=序号, column=10-减列数, value=门诊合计['超封顶金额']).font = font

                ws.cell(row=序号, column=11-减列数, value=门诊合计['自费'])

                nmb = 0
                错误票据标记 = {'负票': [], '重复': [], '票号错': [], '标红': []}   # 记录错误票据的汇总结果
                错误票据筛重 = []
                退单原因 = ''   # 初始化退单原因
                退单汇总 = {}   # 以字典键记录原因，值记录票据数
                for 门诊票据 in 个案信息['stub']:
                    nmb += 1
                    票据信息 = GD.提取案件详情城镇门诊信息(门诊票据)
                    if 基础信息.get('医保类型') == '城镇居民':
                        票据信息['票据类型'] = 1

                    if 票据信息['合计'] <= 0:
                        if nmb not in 错误票据筛重:
                            错误票据标记['负票'].append(f'{nmb}')
                            错误票据筛重.append(nmb)

                    if 票据信息['票据号'] in 票据号:
                        if nmb not in 错误票据筛重:
                            错误票据标记['重复'].append(f'{nmb}')
                            错误票据筛重.append(nmb)
                    
                    if len(票据信息['票据号']) < 6 or 票据信息['票据号'] == 'unknown':
                        if nmb not in 错误票据筛重:
                            错误票据标记['票号错'].append(f'{nmb}')
                            错误票据筛重.append(nmb)

                    if 票据信息['字段标红'] == '是':
                        if nmb not in 错误票据筛重:
                            错误票据标记['标红'].append(f'{nmb}')
                            错误票据筛重.append(nmb)

                    票据号.append(票据信息['票据号'])

                    if 票据年份 == '':
                        票据年份 = 票据信息['票据时间'][2:4]

                    if 票据信息['票据时间'][2:4] not in 票据年份:
                        年份 = 票据信息['票据时间'][2:4]
                        票据年份 = f'{票据年份},{年份}'

                    # 检查家属子女的票是不是大于18岁
                    if '子女' in 基础信息['保单方案']:
                        try:
                            y, m, d = 基础信息['身份证号'][6:10], 基础信息['身份证号'][10:12], 基础信息['身份证号'][12:14]
                            票据日期 = 票据信息['票据时间']
                            birthday = datetime.strptime(f'{y}-{m}-{d}', "%Y-%m-%d")
                            today = datetime.strptime(f'{票据日期}', "%Y-%m-%d")
                            年龄 = (today-birthday).days//365
                            if 年龄 > 17:
                                ws.cell(row=序号, column=19-减列数, value=f'{年龄}岁').fill = fille
                        except:
                            ws.cell(row=序号, column=19-减列数, value='子女身份证号错误').fill = fille

                    if 票据信息['退单状态'] == 1:
                        url = GD.案件详情退单查询网址(票据信息['票据类型'], 票据信息['案件id'], 票据信息['票据id'])
                        data3 = GD.获取案件信息(url, headers)   # 获取退单详情
                        退单内容 = GD.提取退单内容(data3)
                        退单原因 = 退单内容['退单原因']
                        问题描述 = 退单内容['问题描述']
                        # 退单类型 = 退单内容['退单类型'] # 1是整张退单，2是部分退单
                        if 问题描述:
                            退单原因 = f'{退单原因},{问题描述}；'
                        else:
                            退单原因 = f'{退单原因}；'

                        if 退单汇总 == {}:
                            退单汇总 = {退单原因: ['门诊', f'{nmb}']}

                        elif 退单汇总.get(退单原因):
                            退单汇总[退单原因].append(f'{nmb}')
                        
                        else:
                            退单汇总[退单原因] = [f'{nmb}']

                if 错误票据标记['负票'] or 错误票据标记['重复'] or 错误票据标记['票号错'] or 错误票据标记['标红']:
                    错误票据标记 = 字典转文本(错误票据标记)
                    if 错误票据标记:
                        错误标记汇总 += f'门诊:{错误票据标记};'

                if 退单汇总:
                    退单汇总 = 字典转文本(退单汇总)
                    if 退单汇总:
                        退单原因汇总 += 退单汇总

            # 是否有住院
            if 个案信息['stub_hospital']:
                ws.cell(row=序号, column=12-减列数, value='有')
                nmb = 0
                错误票据标记 = {'负票': [], '重复': [], '票号错': [], '标红': []}   # 记录错误票据的汇总结果
                错误票据筛重 = []
                退单汇总 = {}
                for 住院票据 in 个案信息['stub_hospital']:
                    nmb += 1
                    票据信息 = GD.提取案件详情住院信息(住院票据)

                    if 票据信息['合计'] <= 0:
                        if nmb not in 错误票据筛重:
                            错误票据标记['负票'].append(f'{nmb}')
                            错误票据筛重.append(nmb)

                    if 票据信息['票据号'] in 票据号:
                        if nmb not in 错误票据筛重:
                            错误票据标记['重复'].append(f'{nmb}')
                            错误票据筛重.append(nmb)
                    
                    if len(票据信息['票据号']) < 6 or 票据信息['票据号'] == 'unknown':
                        if nmb not in 错误票据筛重:
                            错误票据标记['票号错'].append(f'{nmb}')
                            错误票据筛重.append(nmb)

                    if 票据信息['字段标红'] == '是':
                        if nmb not in 错误票据筛重:
                            错误票据标记['标红'].append(f'{nmb}')
                            错误票据筛重.append(nmb)

                    票据号.append(票据信息['票据号'])

                    if 单位 != '地铁':
                        if 票据年份 == '':
                            票据年份 = 票据信息['入院时间'][2:4]
                            
                        if 票据信息['入院时间'][2:4] not in 票据年份:
                            年份 = 票据信息['入院时间'][2:4]
                            票据年份 = f'{票据年份},{年份}'
                    else:
                        if 票据年份 == '':
                            票据年份 = 票据信息['出院时间'][2:4]
                            
                        if 票据信息['出院时间'][2:4] not in 票据年份:
                            年份 = 票据信息['出院时间'][2:4]
                            票据年份 = f'{票据年份},{年份}'

                        # 检查家属子女的票是不是大于18岁
                        if '子女' in 基础信息['保单方案']:
                            try:
                                y, m, d = 基础信息['身份证号'][6:10], 基础信息['身份证号'][10:12], 基础信息['身份证号'][12:14]
                                票据日期 = 票据信息['出院时间']
                                birthday = datetime.strptime(f'{y}-{m}-{d}', "%Y-%m-%d")
                                today = datetime.strptime(f'{票据日期}', "%Y-%m-%d")
                                年龄 = (today-birthday).days//365
                                if 年龄 > 17:
                                    ws.cell(row=序号, column=19-减列数, value=f'{年龄}岁').fill = fille
                            except:
                                pass
                        
                    if 票据信息['退单状态'] == 1:
                        url = GD.案件详情退单查询网址(票据信息['票据类型'], 票据信息['案件id'], 票据信息['票据id'])
                        data3 = GD.获取案件信息(url, headers)   # 获取退单详情
                        退单内容 = GD.提取退单内容(data3)
                        退单原因 = 退单内容['退单原因']
                        问题描述 = 退单内容['问题描述']
                        # 退单类型 = 退单内容['退单类型']
                        if 问题描述:
                            退单原因 = f'{退单原因},{问题描述}；'
                        else:
                            退单原因 = f'{退单原因}；'

                        if 退单汇总 == {}:
                            退单汇总 = {退单原因: ['住院', f'{nmb}']}

                        elif 退单汇总.get(退单原因):
                            退单汇总[退单原因].append(f'{nmb}')
                        
                        else:
                            退单汇总[退单原因] = [f'{nmb}']

                if 错误票据标记['负票'] or 错误票据标记['重复'] or 错误票据标记['票号错'] or 错误票据标记['标红']:
                    错误票据标记 = 字典转文本(错误票据标记)
                    if 错误票据标记:
                        错误标记汇总 += f'住院:{错误票据标记};'

                if 退单汇总:
                    退单汇总 = 字典转文本(退单汇总)
                    if 退单汇总:
                        退单原因汇总 += 退单汇总

            # 是否有门特
            if 个案信息['stub_whole']:
                ws.cell(row=序号, column=13-减列数, value='有').font = font
                nmb = 0
                错误票据标记 = {'负票': [], '重复': [], '票号错': [], '标红': []}   # 记录错误票据的汇总结果
                错误票据筛重 = []
                退单汇总 = {}
                for 门特票据 in 个案信息['stub_whole']:
                    nmb += 1
                    票据信息 = GD.提取案件详情门特信息(门特票据)

                    if 票据信息['合计'] <= 0:
                        if nmb not in 错误票据筛重:
                            错误票据标记['负票'].append(f'{nmb}')
                            错误票据筛重.append(nmb)

                    if 票据信息['票据号'] in 票据号:
                        if nmb not in 错误票据筛重:
                            错误票据标记['重复'].append(f'{nmb}')
                            错误票据筛重.append(nmb)
                    
                    if len(票据信息['票据号']) < 6 or 票据信息['票据号'] == 'unknown':
                        if nmb not in 错误票据筛重:
                            错误票据标记['票号错'].append(f'{nmb}')
                            错误票据筛重.append(nmb)

                    if 票据信息['字段标红'] == '是':
                        if nmb not in 错误票据筛重:
                            错误票据标记['标红'].append(f'{nmb}')
                            错误票据筛重.append(nmb)

                    票据号.append(票据信息['票据号'])
                            
                    if 票据年份 == '':
                        票据年份 = 票据信息['票据时间'][2:4]

                    if 票据信息['票据时间'][2:4] not in 票据年份:
                        年份 = 票据信息['票据时间'][2:4]
                        票据年份 = f'{票据年份},{年份}'

                    # 检查家属子女的票是不是大于18岁
                    if '子女' in 基础信息['保单方案']:
                        try:
                            y, m, d = 基础信息['身份证号'][6:10], 基础信息['身份证号'][10:12], 基础信息['身份证号'][12:14]
                            票据日期 = 票据信息['票据时间']
                            birthday = datetime.strptime(f'{y}-{m}-{d}', "%Y-%m-%d")
                            today = datetime.strptime(f'{票据日期}', "%Y-%m-%d")
                            年龄 = (today-birthday).days//365
                            if 年龄 > 17:
                                ws.cell(row=序号, column=19-减列数, value=f'{年龄}岁').fill = fille
                        except:
                            ws.cell(row=序号, column=19-减列数, value='子女身份证号错误').fill = fille

                    if 票据信息['退单状态'] == 1:
                        url = GD.案件详情退单查询网址(票据信息['票据类型'], 票据信息['案件id'], 票据信息['票据id'])
                        data3 = GD.获取案件信息(url, headers)   # 获取退单详情
                        退单内容 = GD.提取退单内容(data3)
                        退单原因 = 退单内容['退单原因']
                        问题描述 = 退单内容['问题描述']
                        # 退单类型 = 退单内容['退单类型']
                        if 问题描述:
                            退单原因 = f'{退单原因},{问题描述}；'
                        else:
                            退单原因 = f'{退单原因}；'

                        if 退单汇总 == {}:
                            退单汇总 = {退单原因: ['门特', f'{nmb}']}

                        elif 退单汇总.get(退单原因):
                            退单汇总[退单原因].append(f'{nmb}')
                        
                        else:
                            退单汇总[退单原因] = [f'{nmb}']

                if 错误票据标记['负票'] or 错误票据标记['重复'] or 错误票据标记['票号错'] or 错误票据标记['标红']:
                    错误票据标记 = 字典转文本(错误票据标记)
                    if 错误票据标记:
                        错误标记汇总 += f'门特:{错误票据标记};'

                if 退单汇总:
                    退单汇总 = 字典转文本(退单汇总)
                    if 退单汇总:
                        退单原因汇总 += 退单汇总

            if '未核查' in 错误标记汇总:
                ws.cell(row=序号, column=14-减列数, value=错误标记汇总).fill = fille
            else:
                ws.cell(row=序号, column=14-减列数, value=错误标记汇总)
                
            ws.cell(row=序号, column=15-减列数, value=退单原因汇总)
            ws.cell(row=序号, column=16-减列数, value=票据年份)

        wb.save(f'缓存文件夹/{批次号}.xlsx')
        下载单个文件(批次号)
    else:
        put_text(f'批次{批次号}查询错误！！！！！')

def 检查筛重文件(onefile):
    wb = load_workbook(onefile)
    ws = wb.active      # 获取活跃sheet表
    put_text(f'共{ws.max_row-1}条数据\n')
    nmb = 0
    空行 = 0
    ws['O1'] = '检查状态'

    put_processbar('票据', auto_close=True)
    for row in ws.rows:
        time.sleep(0.1)
        nmb += 1
        set_processbar('票据', nmb / ws.max_row, label=nmb)
        # 检查空行
        if nmb == 1:
            continue
        c = 0
        for cell in row:
            if cell.value is not None:
                c = 1
        if c == 0:
            空行 += 1
            continue
        ws[f'O{nmb}'] = '有重复'

        历史退票, 历史问题件, 票据类型, 历史案件号, 案件号, 案件票据号 = row[5].value, row[6].value, row[3].value, row[7].value, row[11].value, row[12].value
        try:
            # 合计，时间，自付一
            票据校对信息 = [float(row[8].value), row[9].value, float(row[10].value)]
        except:
            ws[f'O{nmb}'] = '无基础数据'
            continue

        if 历史退票 == '是':
            ws[f'O{nmb}'] = '已退单'
            continue
        
        if 历史问题件 == '是':
            ws[f'O{nmb}'] = '历史问题件'
            continue

        # 获取案件信息
        url = GD.案件号查询网址(历史案件号)
        data1_l = GD.获取案件信息(url, headers)
        if not data1_l:
            ws[f'O{nmb}'] = '无历史案件'
            continue

        data2_l = data1_l['data'][0]
        历史个案列表 = GD.提取案件列表个案详情(data2_l)                    
        if 历史个案列表['理算状态'] == '未理算':
            ws[f'O{nmb}'] = '历史未理算'
            continue
        
        # 获取基础信息
        url = GD.案件详情查询网址(历史个案列表['案件id'])
        历史个案信息 = GD.获取案件信息(url, headers)   # 获取个案详情
        历史基础信息 = GD.提取案件详情基础信息(历史个案信息)

        # 检测历史票据明细内容;
        if 票据类型 == '门诊':
            if 历史个案信息['stub']:
                for 门诊票据 in 历史个案信息['stub']:
                    票据信息 = GD.提取案件详情城镇门诊信息(门诊票据)
                    if 历史基础信息.get('医保类型') == '城镇居民':
                        票据信息['票据类型'] = 1
                    # 合计，时间，自付一
                    ls = [票据信息['合计'], 票据信息['票据时间'], 票据信息['自付一']]
                    if 票据校对信息 == ls:
                        if 票据信息['退单状态'] == 1:
                            ws[f'O{nmb}'] = '历史退单'
                            break
                        if 'BJSB' != 历史基础信息['案件号'][:4]:
                            if 票据信息['票据号'] != 案件票据号:
                                ws[f'O{nmb}'] = '票据号不重复'
                                break

        elif 票据类型 == '住院':
            if 历史个案信息['stub_hospital']:
                for 住院票据 in 历史个案信息['stub_hospital']:
                    票据信息 = GD.提取案件详情住院信息(住院票据)
                    # 合计，时间，自付一
                    ls = [票据信息['合计'], 票据信息['出院时间'], 票据信息['自付一']]
                    if 票据校对信息 == ls:
                        if 票据信息['退单状态'] == 1:
                            ws[f'O{nmb}'] = '历史退单'
                            break
                        if 'BJSB' != 历史基础信息['案件号'][:4]:
                            if 票据信息['票据号'] != 案件票据号:
                                ws[f'O{nmb}'] = '票据号不重复'
                                break

        elif 票据类型 == '门特':
            if 历史个案信息['stub_whole']:
                for 门特票据 in 历史个案信息['stub_whole']:
                    票据信息 = GD.提取案件详情门特信息(门特票据)
                    # 合计，时间，自付一
                    ls = [票据信息['合计'], 票据信息['票据时间'], 票据信息['自付一']]
                    if 票据校对信息 == ls:
                        if 票据信息['退单状态'] == 1:
                            ws[f'O{nmb}'] = '历史退单'
                            break
                        if 'BJSB' != 历史基础信息['案件号'][:4]:
                            if 票据信息['票据号'] != 案件票据号:
                                ws[f'O{nmb}'] = '票据号不重复'
                                break
        
        url = GD.案件号查询网址(案件号)
        data1 = GD.获取案件信息(url, headers)
        data2 = data1['data'][0]
        个案列表 = GD.提取案件列表个案详情(data2)
        # 获取基础信息
        url = GD.案件详情查询网址(个案列表['案件id'])
        个案信息 = GD.获取案件信息(url, headers)   # 获取个案详情
        基础信息 = GD.提取案件详情基础信息(个案信息)

        if 基础信息['问题件'] == '是':
            ws[f'O{nmb}'] = '本案问题件'
            continue

        # 检测本案票据明细内容;
        if 票据类型 == '门诊':
            if 个案信息['stub']:
                for 门诊票据 in 个案信息['stub']:
                    状态 = 0
                    票据信息 = GD.提取案件详情城镇门诊信息(门诊票据)
                    if 基础信息.get('医保类型') == '城镇居民':
                        票据信息['票据类型'] = 1
                    # 合计，时间，自付一
                    ls = [票据信息['合计'], 票据信息['票据时间'], 票据信息['自付一']]
                    if 票据校对信息 == ls:
                        if 票据信息['退单状态'] == 1:
                            ws[f'O{nmb}'] = '本案退单'
                            状态 = 1
                            break
                if 状态 == 1:
                    continue

        elif 票据类型 == '住院':
            if 个案信息['stub_hospital']:
                for 住院票据 in 个案信息['stub_hospital']:
                    状态 = 0
                    票据信息 = GD.提取案件详情住院信息(住院票据)
                    # 合计，时间，自付一
                    ls = [票据信息['合计'], 票据信息['出院时间'], 票据信息['自付一']]
                    if 票据校对信息 == ls:
                        if 票据信息['退单状态'] == 1:
                            ws[f'O{nmb}'] = '本案退单'
                            状态 = 1
                            break
                if 状态 == 1:
                    continue

        elif 票据类型 == '门特':
            if 个案信息['stub_whole']:
                for 门特票据 in 个案信息['stub_whole']:
                    状态 = 0
                    票据信息 = GD.提取案件详情门特信息(门特票据)
                    # 合计，时间，自付一
                    ls = [票据信息['合计'], 票据信息['票据时间'], 票据信息['自付一']]
                    if 票据校对信息 == ls:
                        if 票据信息['退单状态'] == 1:
                            ws[f'O{nmb}'] = '本案退单'
                            状态 = 1
                            break
                if 状态 == 1:
                    continue

    wb.save(onefile)

def 批次号导出票据明细表(批次号, 单位, 单位简称=''):
    today = date.today()
    lst = []
    lstw = []
    # 获取案件信息
    url = GD.批次号查询网址(批次号)
    data1 = GD.获取案件信息(url, headers)

    if data1 == '没有更多啦~':
        put_text(f'系统没有查询到此批次号{批次号}')
        lst.append([])
        lstw.append(['', 批次号, '', '没有查询到此批次号'])
        return lst, lstw

    案件总数 = data1['page']['count']
    案件总页数 = data1['page']['pages']
    put_text(f'批次{批次号}共{案件总数}件案件。\n')
    data2 = data1['data']
    if 案件总页数 > 1:
        # 按页循环获取数据
        for page in range(2, 案件总页数+1):
            url = GD.批次号查询网址(批次号, 10, page)
            data1 = GD.获取案件信息(url, headers)   # 获取案件列表页
            # 把每页的数据合并到一起
            for i in data1['data']:
                data2.append(i)

    n = 0
    put_processbar(批次号, auto_close=True)
    for x in data2:
        time.sleep(0.1)
        n += 1
        lst1 = []
        set_processbar(批次号, n / 案件总数, label=n)
        个案列表 = GD.提取案件列表个案详情(x)
        上传时间 = 个案列表['上传时间'][:7]

        # 获取基础信息
        url = GD.案件详情查询网址(个案列表['案件id'])
        个案信息 = GD.获取案件信息(url, headers)   # 获取个案详情
        if 个案信息:
            基础信息 = GD.提取案件详情基础信息(个案信息)
        else:
            l = [基础信息['姓名'], 基础信息['身份证号'], 基础信息['案件号'], '读取案件详情时出错']
            lstw.append(l)
            continue

        if 基础信息['问题件'] == '是':  # 问题件跳过
            l = [基础信息['姓名'], 基础信息['身份证号'], 基础信息['案件号'], '问题件', 基础信息['问题件简述'], 基础信息['备注']]
            lstw.append(l)
            continue

        # 检测票据明细内容;
        # 是否有门诊
        if 个案信息['stub']:
            for 门诊票据 in 个案信息['stub']:
                
                票据信息 = GD.提取案件详情城镇门诊信息(门诊票据)
                if 基础信息.get('医保类型') == '城镇居民':
                    票据信息['票据类型'] = 1

                # '票据号', '自付一', '起付金额', '超封顶金额', '自付二', '自费', '合计', '票据时间', '出院时间', '票据类型', '备注'
                if 票据信息['退单状态'] == 1:
                    continue

                # 添加基础信息
                if 单位 == '公交':
                    lst1 = [基础信息['姓名'], 基础信息['性别'], 基础信息['身份证号'], 基础信息['案件号']]
                else:
                    lst1 = [基础信息['姓名'], 基础信息['身份证号'], 基础信息['案件号']]

                lst1.append(票据信息['票据号'])
                lst1.append(票据信息['自付一'])
                lst1.append(票据信息['起付金额'])
                lst1.append(票据信息['超封顶金额'])
                lst1.append(票据信息['自付二'])
                lst1.append(票据信息['自费'])
                lst1.append(票据信息['个人支付'])
                lst1.append(票据信息['票据时间'])
                lst1.append('')
                lst1.append('普通门诊')
                lst1.append(基础信息['备注'])
                if 单位 == '地铁':
                    lst1.append(上传时间)
                else:
                    lst1.append(单位简称)

                lst.append(lst1)

        # 是否有住院
        if 个案信息['stub_hospital']:
            for 住院票据 in 个案信息['stub_hospital']:
                票据信息 = GD.提取案件详情住院信息(住院票据)

                # '票据号', '自付一', '起付金额', '超封顶金额', '自付二', '自费', '合计', '票据时间', '出院时间', '票据类型', '备注'
                if 票据信息['退单状态'] == 1:
                    continue

                # 添加基础信息
                if 单位 == '公交':
                    lst1 = [基础信息['姓名'], 基础信息['性别'], 基础信息['身份证号'], 基础信息['案件号']]
                else:
                    lst1 = [基础信息['姓名'], 基础信息['身份证号'], 基础信息['案件号']]

                lst1.append(票据信息['票据号'])
                lst1.append(票据信息['自付一'])
                lst1.append(票据信息['起付金额'])
                lst1.append(票据信息['超封顶金额'])
                lst1.append(票据信息['自付二'])
                lst1.append(票据信息['自费'])
                lst1.append(票据信息['个人支付'])
                lst1.append(票据信息['入院时间'])
                lst1.append(票据信息['出院时间'])
                lst1.append('普通住院')
                lst1.append(基础信息['备注'])
                if 单位 == '地铁':
                    lst1.append(上传时间)
                else:
                    lst1.append(单位简称)

                lst.append(lst1)

        # 是否有门特
        if 个案信息['stub_whole']:
            for 门特票据 in 个案信息['stub_whole']:
                票据信息 = GD.提取案件详情门特信息(门特票据)

                # '票据号', '自付一', '起付金额', '超封顶金额', '自付二', '自费', '合计', '票据时间', '出院时间', '票据类型', '备注'
                if 票据信息['退单状态'] == 1:
                    continue
                # 添加基础信息
                if 单位 == '公交':
                    lst1 = [基础信息['姓名'], 基础信息['性别'], 基础信息['身份证号'], 基础信息['案件号']]
                else:
                    lst1 = [基础信息['姓名'], 基础信息['身份证号'], 基础信息['案件号']]

                lst1.append(票据信息['票据号'])
                lst1.append(票据信息['自付一'])
                lst1.append(票据信息['起付金额'])
                lst1.append(票据信息['超封顶金额'])
                lst1.append(票据信息['自付二'])
                lst1.append(票据信息['自费'])
                lst1.append(票据信息['个人支付'])
                lst1.append(票据信息['票据时间'])

                if 单位 == '地铁':
                    lst1.append(票据信息['票据时间'])
                else:
                    lst1.append('')

                lst1.append('特殊门诊')
                lst1.append(基础信息['备注'])
                if 单位 == '地铁':
                    lst1.append(上传时间)
                else:
                    lst1.append(单位简称)

                lst.append(lst1)

        if not lst:
            lstw.append([基础信息['姓名'], 基础信息['身份证号'], 基础信息['案件号'], '无票据'])

    return lst, lstw

def 批次号导出理算结果表(批次号, 单位, 单位简称=''):
    lst = []
    lstw = []
    # 获取案件信息
    url = GD.批次号查询网址(批次号)
    data1 = GD.获取案件信息(url, headers)

    if data1 == '没有更多啦~':
        put_text(f'系统没有查询到此批次号{批次号}')
        lst.append([])
        lstw.append(['', 批次号, '', '没有查询到此批次号'])
        return lst, lstw

    案件总数 = data1['page']['count']
    案件总页数 = data1['page']['pages']
    put_text(f'批次{批次号}共{案件总数}件案件。\n')
    data2 = data1['data']
    if 案件总页数 > 1:
        # 按页循环获取数据
        for page in range(2, 案件总页数+1):
            url = GD.批次号查询网址(批次号, 10, page)
            data1 = GD.获取案件信息(url, headers)   # 获取案件列表页
            # 把每页的数据合并到一起
            for i in data1['data']:
                data2.append(i)

    n = 0
    put_processbar(批次号, auto_close=True)
    for x in data2:
        time.sleep(0.1)
        n += 1
        set_processbar(批次号, n / 案件总数, label=n)
        个案列表 = GD.提取案件列表个案详情(x)
        票据年份 = ''

        if 个案列表['理算状态'] != '已理算':
            # 获取基础信息
            url = GD.案件详情查询网址(个案列表['案件id'])
            个案信息 = GD.获取案件信息(url, headers)   # 获取个案详情
            if 个案信息:
                基础信息 = GD.提取案件详情基础信息(个案信息)
            else:
                案件号 = 基础信息['案件号']
                l = [基础信息['姓名'], 基础信息['身份证号'], 基础信息['案件号'], '读取案件详情时出错']
                lstw.append(l)
                put_text(f'{案件号}导出错误!!!!!!!!!')
                continue

            if 基础信息['问题件'] == '是':  # 问题件跳过
                l = [基础信息['姓名'], 基础信息['身份证号'], 基础信息['案件号'], '问题件']
                lstw.append(l)
                continue

            l = [个案列表['姓名'], 个案列表['身份证号'], 个案列表['案件号'], '未理算']
            lstw.append(l)
            continue

        # 获取基础信息
        url = GD.案件详情查询网址(个案列表['案件id'])
        个案信息 = GD.获取案件信息(url, headers)   # 获取个案详情
        if 个案信息:
            基础信息 = GD.提取案件详情基础信息(个案信息)
        else:
            案件号 = 基础信息['案件号']
            l = [基础信息['姓名'], 基础信息['身份证号'], 基础信息['案件号'], '读取案件详情时出错']
            lstw.append(l)
            put_text(f'{案件号}导出错误!!!!!!!!!')
            continue

        if '公交' in 单位:
            lst1 = [基础信息['姓名'], 基础信息['性别'], 基础信息['身份证号'], 基础信息['案件号']]
        else:
            lst1 = [基础信息['姓名'], 基础信息['身份证号'], 基础信息['案件号']]

        # 检测票据明细内容;
        # 是否有门诊
        if 个案信息['stub']:
            门诊合计 = GD.提取案件详情城镇门诊合计信息(个案信息['sum_stub'])
            for 门诊信息 in 个案信息['stub']:
                票据信息 = GD.提取案件详情城镇门诊信息(门诊信息)
                if 基础信息.get('医保类型') == '城镇居民':
                    票据信息['票据类型'] = 1
                if 票据信息['退单状态'] == 0:
                    票据年份 = 票据信息['票据时间'][2:4]
                    break

            lst1.append(门诊合计['自付一'])
            lst1.append(门诊合计['起付金额'])
            lst1.append(门诊合计['超封顶金额'])
            lst1.append(门诊合计['自付二'])
            lst1.append(门诊合计['自费'])

        else:
            lst1.append(0)
            lst1.append(0)
            lst1.append(0)
            lst1.append(0)
            lst1.append(0)

        # 是否有住院
        if 个案信息['stub_hospital']:
            住院合计 = GD.提取案件详情住院合计信息(个案信息['sum_stub_hospital'])
            if not 票据年份:
                for 住院信息 in 个案信息['stub_hospital']:
                    票据信息 = GD.提取案件详情住院信息(住院信息)
                    if 票据信息['退单状态'] == 0:
                        if 单位 == "地铁":
                            票据年份 = 票据信息['出院时间'][2:4]
                            break
                        else:
                            票据年份 = 票据信息['入院时间'][2:4]
                            break
            lst1.append(住院合计['自付一'])
            lst1.append(住院合计['起付金额'])
            lst1.append(住院合计['超封顶金额'])
            lst1.append(住院合计['自付二'])
            lst1.append(住院合计['自费'])
        else:
            lst1.append(0)
            lst1.append(0)
            lst1.append(0)
            lst1.append(0)
            lst1.append(0)
                
        # 是否有门特
        if 个案信息['stub_whole']:
            门特合计 = GD.提取案件详情门特合计信息(个案信息['sum_stub_whole'])
            if not 票据年份:
                for 门特信息 in 个案信息['stub_whole']:
                    票据信息 = GD.提取案件详情门特信息(门特信息)
                    if 票据信息['退单状态'] == 0:
                        票据年份 = 票据信息['票据时间'][2:4]
                        break
            lst1.append(门特合计['自付一'])
            lst1.append(门特合计['起付金额'])
            lst1.append(门特合计['超封顶金额'])
            lst1.append(门特合计['自付二'])
            lst1.append(门特合计['自费'])
        else:
            lst1.append(0)
            lst1.append(0)
            lst1.append(0)
            lst1.append(0)
            lst1.append(0)
                
        try:
            # 获取理算结果
            url = GD.列表页理算结果查询网址(个案列表['案件号'])
            理算信息 = GD.获取案件信息(url, headers)
            理算结果 = GD.提取列案件表页理算结果(理算信息)

            lst1.append(理算结果['门诊回传金额'])
            lst1.append(理算结果['住院回传总额'])
            lst1.append(理算结果['回传总额'])

        except Exception as e:
            # 输出错误提示
            print(datetime.now())
            print(traceback.format_exc())
            print('====='*50)
            print(e)
            l = [个案列表['姓名'], 个案列表['身份证号'], 个案列表['案件号'], '读取理算结果时出错']
            lstw.append(l)
            # 输出错误提示
            continue
        
        lst1.append(票据年份)
        if 单位 == '公交':
            lst1.append(单位简称)

        lst.append(lst1)
    
    return lst, lstw

def 案件号导出票据明细表(案件号, 单位, 单位简称=''):
    # today = date.today()
    lst = []
    lst1 = []
    # 获取案件信息
    url = GD.案件号查询网址(案件号)
    data1 = GD.获取案件信息(url, headers)
    
    if data1 == '没有更多啦~':
        lst.append(['', 案件号, '系统没有此案件'])
        return lst

    data2 = data1['data']

    for x in data2:
        time.sleep(0.1)
        个案列表 = GD.提取案件列表个案详情(x)
        上传时间 = 个案列表['上传时间'][:7]

        # 获取基础信息
        url = GD.案件详情查询网址(个案列表['案件id'])
        个案信息 = GD.获取案件信息(url, headers)   # 获取个案详情
        if 个案信息:
            基础信息 = GD.提取案件详情基础信息(个案信息)
        else:
            案件号 = 个案列表['案件号']
            # print(f'{案件号}没有票据明细')
            put_text(f'{案件号}错误！！！！！')
            continue

        if 基础信息['问题件'] == '是':  # 问题件跳过
            put_text(f'{案件号}是问题件！！！！！')
            continue

        # 检测票据明细内容;
        # 是否有门诊
        if 个案信息['stub']:
            for 门诊票据 in 个案信息['stub']:
                票据信息 = GD.提取案件详情城镇门诊信息(门诊票据)
                if 基础信息.get('医保类型') == '城镇居民':
                    票据信息['票据类型'] = 1

                # '票据号', '自付一', '起付金额', '超封顶金额', '自付二', '自费', '合计', '票据时间', '出院时间', '票据类型', '备注'
                if 票据信息['退单状态'] == 1:
                    continue
                # 添加基础信息
                if 单位 == '公交':
                    lst1 = [基础信息['姓名'], 基础信息['性别'], 基础信息['身份证号'], 基础信息['案件号']]
                else:
                    lst1 = [基础信息['姓名'], 基础信息['身份证号'], 基础信息['案件号']]

                lst1.append(票据信息['票据号'])
                lst1.append(票据信息['自付一'])
                lst1.append(票据信息['起付金额'])
                lst1.append(票据信息['超封顶金额'])
                lst1.append(票据信息['自付二'])
                lst1.append(票据信息['自费'])
                lst1.append(票据信息['个人支付'])
                lst1.append(票据信息['票据时间'])
                lst1.append('')
                lst1.append('普通门诊')
                lst1.append(基础信息['备注'])
                if 单位 == '地铁':
                    lst1.append(上传时间)
                else:
                    lst1.append(单位简称)

                lst.append(lst1)

        # 是否有住院
        if 个案信息['stub_hospital']:
            for 住院票据 in 个案信息['stub_hospital']:
                票据信息 = GD.提取案件详情住院信息(住院票据)

                # '票据号', '自付一', '起付金额', '超封顶金额', '自付二', '自费', '合计', '票据时间', '出院时间', '票据类型', '备注'
                if 票据信息['退单状态'] == 1:
                    continue
                # 添加基础信息
                if 单位 == '公交':
                    lst1 = [基础信息['姓名'], 基础信息['性别'], 基础信息['身份证号'], 基础信息['案件号']]
                else:
                    lst1 = [基础信息['姓名'], 基础信息['身份证号'], 基础信息['案件号']]

                lst1.append(票据信息['票据号'])
                lst1.append(票据信息['自付一'])
                lst1.append(票据信息['起付金额'])
                lst1.append(票据信息['超封顶金额'])
                lst1.append(票据信息['自付二'])
                lst1.append(票据信息['自费'])
                lst1.append(票据信息['个人支付'])
                lst1.append(票据信息['入院时间'])
                lst1.append(票据信息['出院时间'])
                lst1.append('普通住院')
                lst1.append(基础信息['备注'])
                if 单位 == '地铁':
                    lst1.append(上传时间)
                else:
                    lst1.append(单位简称)

                lst.append(lst1)

        # 是否有门特
        if 个案信息['stub_whole']:
            for 门特票据 in 个案信息['stub_whole']:
                票据信息 = GD.提取案件详情门特信息(门特票据)

                # '票据号', '自付一', '起付金额', '超封顶金额', '自付二', '自费', '合计', '票据时间', '出院时间', '票据类型', '备注'
                if 票据信息['退单状态'] == 1:
                    continue
                # 添加基础信息
                if 单位 == '公交':
                    lst1 = [基础信息['姓名'], 基础信息['性别'], 基础信息['身份证号'], 基础信息['案件号']]
                else:
                    lst1 = [基础信息['姓名'], 基础信息['身份证号'], 基础信息['案件号']]

                lst1.append(票据信息['票据号'])
                lst1.append(票据信息['自付一'])
                lst1.append(票据信息['起付金额'])
                lst1.append(票据信息['超封顶金额'])
                lst1.append(票据信息['自付二'])
                lst1.append(票据信息['自费'])
                lst1.append(票据信息['个人支付'])
                lst1.append(票据信息['票据时间'])
                
                if 单位 == '地铁':
                    lst1.append(票据信息['票据时间'])
                else:
                    lst1.append('')

                lst1.append('特殊门诊')
                lst1.append(基础信息['备注'])
                if 单位 == '地铁':
                    lst1.append(上传时间)
                else:
                    lst1.append(单位简称)

                lst.append(lst1)
        if not lst:
            lst.append(['无票据'])

    return lst

def 案件号导出理算结果表(案件号, 单位, 单位简称=''):
    lst = []
    # 获取案件信息
    url = GD.案件号查询网址(案件号)
    data1 = GD.获取案件信息(url, headers)
    
    if data1 == '没有更多啦~':
        lst.append(['', 案件号, '系统没有此案件'])
        return lst

    data2 = data1['data']

    for x in data2:
        time.sleep(0.1)
        个案列表 = GD.提取案件列表个案详情(x)
        票据年份 = ''

        if 个案列表['理算状态'] != '已理算':
            # put_text(f'{案件号}没有理算结果！！！！！')
            lst.append(['', '', 案件号, '未理算'])
            return lst
        
        # 获取基础信息
        url = GD.案件详情查询网址(个案列表['案件id'])
        个案信息 = GD.获取案件信息(url, headers)   # 获取个案详情
        if 个案信息:
            基础信息 = GD.提取案件详情基础信息(个案信息)
        else:
            案件号 = 个案列表['案件号']
            # put_text(f'{案件号}错误！！！！！')
            lst.append(['', '', 案件号, '导出错误'])
            return lst

        if 基础信息['问题件'] == '是':
            lst.append(['', '', 案件号, '问题件'])
            return lst
            
        if 单位 == '地铁':
            lst1 = [基础信息['姓名'], 基础信息['身份证号'], 基础信息['案件号']]
        else:
            lst1 = [基础信息['姓名'], 基础信息['性别'], 基础信息['身份证号'], 基础信息['案件号']]

        # 检测票据明细内容;
        # 是否有门诊
        if 个案信息['stub']:
            门诊合计 = GD.提取案件详情城镇门诊合计信息(个案信息['sum_stub'])
            for 门诊信息 in 个案信息['stub']:
                票据信息 = GD.提取案件详情城镇门诊信息(门诊信息)
                if 基础信息.get('医保类型') == '城镇居民':
                    票据信息['票据类型'] = 1
                if 票据信息['退单状态'] == 0:
                    票据年份 = 票据信息['票据时间'][2:4]
                    break

            lst1.append(门诊合计['自付一'])
            lst1.append(门诊合计['起付金额'])
            lst1.append(门诊合计['超封顶金额'])
            lst1.append(门诊合计['自付二'])
            lst1.append(门诊合计['自费'])

        else:
            lst1.append(0)
            lst1.append(0)
            lst1.append(0)
            lst1.append(0)
            lst1.append(0)

        # 是否有住院
        if 个案信息['stub_hospital']:
            住院合计 = GD.提取案件详情住院合计信息(个案信息['sum_stub_hospital'])
            if not 票据年份:
                for 住院信息 in 个案信息['stub_hospital']:
                    票据信息 = GD.提取案件详情住院信息(住院信息)
                    if 票据信息['退单状态'] == 0:
                        if 单位 == "地铁":
                            票据年份 = 票据信息['出院时间'][2:4]
                            break
                        else:
                            票据年份 = 票据信息['入院时间'][2:4]
                            break
            lst1.append(住院合计['自付一'])
            lst1.append(住院合计['起付金额'])
            lst1.append(住院合计['超封顶金额'])
            lst1.append(住院合计['自付二'])
            lst1.append(住院合计['自费'])
        else:
            lst1.append(0)
            lst1.append(0)
            lst1.append(0)
            lst1.append(0)
            lst1.append(0)
                
        # 是否有门特
        if 个案信息['stub_whole']:
            门特合计 = GD.提取案件详情门特合计信息(个案信息['sum_stub_whole'])
            if not 票据年份:
                for 门特信息 in 个案信息['stub_whole']:
                    票据信息 = GD.提取案件详情门特信息(门特信息)
                    if 票据信息['退单状态'] == 0:
                        票据年份 = 票据信息['票据时间'][2:4]
                        break
            lst1.append(门特合计['自付一'])
            lst1.append(门特合计['起付金额'])
            lst1.append(门特合计['超封顶金额'])
            lst1.append(门特合计['自付二'])
            lst1.append(门特合计['自费'])
        else:
            lst1.append(0)
            lst1.append(0)
            lst1.append(0)
            lst1.append(0)
            lst1.append(0)
                
        try:
            # 获取理算结果
            url = GD.列表页理算结果查询网址(个案列表['案件号'])
            理算信息 = GD.获取案件信息(url, headers)
            理算结果 = GD.提取列案件表页理算结果(理算信息)

            lst1.append(理算结果['门诊回传金额'])
            lst1.append(理算结果['住院回传总额'])
            lst1.append(理算结果['回传总额'])
        
        except Exception as e:
            # 输出错误提示
            print(datetime.now())
            print(traceback.format_exc())
            print('====='*50)
            print(e)
            案件号 = 个案列表['案件号']
            # print(f'{案件号}没有理算结果')
            put_text(f'{案件号}没有理算结果！！！！！')
            lst1.append('没有理算结果')
            continue
        
        lst1.append(票据年份)
        if 单位 == '公交':
            lst1.append(单位简称)

        lst.append(lst1)

    return lst

def 批次号导出案件列表信息(批次号, 选项):
    lst = []
    # 获取案件信息
    url = GD.批次号查询网址(批次号)
    data1 = GD.获取案件信息(url, headers)

    if data1 == '没有更多啦~':
        put_text(f'系统没有查询到{批次号}')
        return data1

    案件总数 = data1['page']['count']
    案件总页数 = data1['page']['pages']
    put_text(f'批次{批次号}共{案件总数}件案件。')
    data2 = data1['data']
    if 案件总页数 > 1:
        # 按页循环获取数据
        for page in range(2, 案件总页数+1):
            url = GD.批次号查询网址(批次号, 10, page)
            data1 = GD.获取案件信息(url, headers)   # 获取案件列表页
            # 把每页的数据合并到一起
            for i in data1['data']:
                data2.append(i)
    n = 0
    put_processbar(批次号, auto_close=True)
    for x in data2:
        time.sleep(0.1)
        n += 1
        set_processbar(批次号, n / 案件总数, label=n)
        个案列表 = GD.提取案件列表个案详情(x)

        序号 = 个案列表['案件号'][len(批次号):]
        if 选项 == '是':
            lst1 = [序号, 个案列表['姓名'], 个案列表['身份证号'], 个案列表['批次号'], 个案列表['案件号'], 个案列表['上传时间'][:10], 个案列表['回传时间'][:10], 个案列表['理算状态']]
        else:
            lst1 = [序号, 个案列表['姓名'], 个案列表['身份证号'], 个案列表['案件号']]

        lst.append(lst1)

    return lst

def 案件号导出案件列表信息(案件号, 序号):
    # 获取案件信息
    url = GD.案件号查询网址(案件号)
    data1 = GD.获取案件信息(url, headers)
    
    if data1 == '没有更多啦~':
        lst.append([序号, 案件号, '系统没有此案件'])
        return lst

    data2 = data1['data']

    个案列表 = GD.提取案件列表个案详情(data2[0])

    lst = [序号, 个案列表['姓名'], 个案列表['身份证号'], 个案列表['批次号'], 个案列表['案件号'], 个案列表['上传时间'][:10], 个案列表['回传时间'][:10], 个案列表['理算状态']]

    return lst

def 个人身份证影像查询(身份证号):
    lst = []
    # 获取案件信息
    url = GD.身份证号查询网址(身份证号)
    data1 = GD.获取案件信息(url, headers)
    
    if data1 == '没有更多啦~':
        lst.append(['', 身份证号, '系统没有此人'])
        return lst

    案件总数 = data1['page']['count']
    if 案件总数 > 10:   # 如果数量大于1页，重新获取全部案件
        url = GD.身份证号查询网址(身份证号, 案件总数)
        data1 = GD.获取案件信息(url, headers)   # 获取案件列表页
    data2 = data1['data']

    nmb = 1
    put_processbar('案件号', auto_close=True)
    for x in data2:
        set_processbar('案件号', nmb / len(data2), label=nmb)
        nmb += 1
        time.sleep(0.1)
        个案列表 = GD.提取案件列表个案详情(x)
        # 获取基础信息
        url = GD.案件详情查询网址(个案列表['案件id'])
        个案信息 = GD.获取案件信息(url, headers)   # 获取个案详情
        if 个案信息:
            基础信息 = GD.提取案件详情基础信息(个案信息)
        else:
            案件号 = 个案列表['案件号']
            put_text(f'{案件号}错误！！！！！')
            continue

        # 获取基础信息并查询身份证栏
        lst1 = [基础信息['姓名'], 基础信息['身份证号']]
        # 查询身份证栏
        if 个案信息['stub_id_card']:
            lst1.append(基础信息['案件号'])
            nmb = 案件总数
            set_processbar('案件号', nmb / len(data2), label=nmb)
            break
    lst.append(lst1)
    return lst

def 身份证号指定条件查询案件号(身份证号, 查询条件):
    lst = []
    # 获取案件信息
    url = GD.身份证号查询网址(身份证号)
    data1 = GD.获取案件信息(url, headers)
    
    if data1 == '没有更多啦~':
        lst.append(['', 身份证号, '系统没有此人'])
        return lst
    案件总数 = data1['page']['count']
    if 案件总数 > 10:   # 如果数量大于1页，重新获取全部案件
        url = GD.身份证号查询网址(身份证号, 案件总数)
        data1 = GD.获取案件信息(url, headers)   # 获取案件列表页
    data2 = data1['data']

    nmb = 1
    put_processbar('案件号', auto_close=True)
    for x in data2:
        set_processbar('案件号', nmb / len(data2), label=nmb)
        nmb += 1
        time.sleep(0.1)
        个案列表 = GD.提取案件列表个案详情(x)

        if 查询条件:
            if 查询条件 in 个案列表['上传时间'][:7]:
                lst.append([个案列表['姓名'], 个案列表['身份证号'], 个案列表['案件号'], 个案列表['上传时间'][:10], 个案列表['理算状态']])
            else:
                continue
        else:
            lst.append([个案列表['姓名'], 个案列表['身份证号'], 个案列表['案件号'], 个案列表['上传时间'][:10], 个案列表['理算状态']])

    return lst

# 赔付明细操作
def 地铁全量赔付查询(赔付明细, 身份证号):
    基础信息 = []
    历史赔付20 = []
    历史赔付21 = []
    历史赔付22 = []
    if 赔付明细.get(身份证号):
        姓名 = 赔付明细.get(身份证号).get('姓名')
        
        基础信息.append(['单位', '姓名', '身份证号', '客户号', '原方案', '现方案', '修改日期', '超限额'])
        ls = [
            赔付明细.get(身份证号).get('单位'),
            赔付明细.get(身份证号).get('姓名'),
            身份证号,
            赔付明细.get(身份证号).get('客户号'),
            赔付明细.get(身份证号).get('原方案'),
            赔付明细.get(身份证号).get('现方案'),
            赔付明细.get(身份证号).get('修改日期'),
            赔付明细.get(身份证号).get('超限额')
        ]
        基础信息.append(ls)

        if 赔付明细.get(身份证号).get('2020'):
            历史赔付20.append(['年份', '年度已报销', '年度未报销', '本次已报销', '本次未报销', '赔付日期'])
            ls = 赔付明细.get(身份证号).get('2020')
            # '年份', '年度已报销':'', '年度未报销':'', '历史报销':[[0'本次已报销', 1'本次未报销', 2'赔付日期']]
            for i in ls.get('历史报销'):
                lst = []
                lst.append('2020')
                lst.append(round(ls.get('年度已报销'), 2))
                lst.append(round(ls.get('年度未报销'), 2))
                for j in i:
                    lst.append(j)

                历史赔付20.append(lst)

        if 赔付明细.get(身份证号).get('2021'):
            历史赔付21.append(['年份', '年度已报销', '年度未报销', '本次已报销', '本次未报销', '赔付日期'])
            ls = 赔付明细.get(身份证号).get('2021')
            for i in ls.get('历史报销'):
                lst = []
                lst.append('2021')
                lst.append(round(ls.get('年度已报销'), 2))
                lst.append(round(ls.get('年度未报销'), 2))
                for j in i:
                    lst.append(j)

                历史赔付21.append(lst)

        if 赔付明细.get(身份证号).get('2022'):
            历史赔付22.append(['年份', '年度已报销', '年度未报销', '本次已报销', '本次未报销', '赔付日期'])
            ls = 赔付明细.get(身份证号).get('2022')
            for i in ls.get('历史报销'):
                lst = []
                lst.append('2022')
                lst.append(round(ls.get('年度已报销'), 2))
                lst.append(round(ls.get('年度未报销'), 2))
                for j in i:
                    lst.append(j)

                历史赔付22.append(lst)
        
        with put_collapse(f'点击查看“{姓名}”赔付详情：'):
            put_table(基础信息)
            if 历史赔付20:
                with put_collapse('点击查看20年赔付明细：'):
                    put_table(历史赔付20)

            if 历史赔付21:
                with put_collapse('点击查看21年赔付明细：'):
                    put_table(历史赔付21)

            if 历史赔付22:
                with put_collapse('点击查看22年赔付明细：'):
                    put_table(历史赔付22)

    else:
        popup('没有此人')

# 下面是退单函数
def 地铁超限额退单表(表格路径, 选项='是'):
    # 样式
    thin = Side(border_style="thin", color="000000")  #边框样式，颜色
    border = Border(left=thin, right=thin, top=thin, bottom=thin)   #边框的位置
    font = Font(size=14, bold=True, name='微软雅黑',  color="FF0000")   #字体大小，加粗，字体名称，字体名字
    fill = PatternFill(patternType="solid", start_color='FBEFF2')  # 填充
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)   # 字体居中并自动换行

    # 开始读取原表格内容
    wb_old = load_workbook(表格路径)
    ws_old = wb_old.active
    put_text(f'共导入{ws_old.max_row-1}条数据')
    nmb = 0
    空行 = 0
    put_processbar('row', auto_close=True)
    for row in ws_old.rows:
        nmb += 1
        set_processbar('row', nmb / ws_old.max_row, label=nmb)

        # 检查空行
        c = 0
        for cell in row:
            if cell.value is not None:
                c = 1
        if c == 0:
            空行 += 1
            continue
        
        if nmb-空行 == 1:
            # 读取当天日期
            dates = datetime.now().strftime('%Y-%m-%d')

            # 创建工作薄
            wb = Workbook()
            ws = wb.active
            ws.title = '退单汇总'

            # 设置首行首列为10做为边界空隙，设置第二行行高为5
            ws.row_dimensions[1].height = 10
            # ws.column_dimensions['A'].width = 2
            # 设置单元格行高和列宽，行高依次为2:60放图片，3:20，4:15，列宽依次为B:3，C:17，D:7，E:20，F:20，G:18，H:6，I:4，J:8
            ws.row_dimensions[2].height = 40
            ws.row_dimensions[3].height = 20
            ws.row_dimensions[4].height = 15
            ws.column_dimensions['A'].width = 3.5
            ws.column_dimensions['B'].width = 17
            ws.column_dimensions['C'].width = 7
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 18
            ws.column_dimensions['G'].width = 8
            ws.column_dimensions['H'].width = 6
            ws.column_dimensions['I'].width = 10

            # 读取图片并添加到工作表并定位到单元格旁边
            img = Image('北京人寿.png')
            ws.add_image(img, 'B2')

            # 合并第3、4行的单元格
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=9)
            ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=9)

            # 写入内容并设置居中和居左
            # 设置粗体
            # 单位名称 = '北京市地铁运营有限公司'
            单位 = f'理赔退单交接表（北京市地铁运营有限公司）'
            ws.cell(row=3, column=1, value=单位).font = Font(size=14, bold=True)
            # 字体居中并循环加入下边框
            ws['A3'].alignment = alignment
            for row in ws['A3:I3']:
                for cell in row:
                    cell.border = Border(bottom=thin)

            # 获取现在的时间并格式化，左对齐
            ws['A4'] = dates
            ws['A4'].alignment = Alignment(horizontal='left', vertical='center')

            # 加入标题
            ws.append(['序号', '单位名称', '姓名', '证件号码', '退单原因', '影像名称', '退单票据份数', '信封编号', '退单类型'])
            # 填充颜色
            for row in ws['A5:I5']:
                for cell in row:
                    # 填充颜色、自动换行、加边框、字体居中
                    cell.fill = fill
                    cell.alignment = alignment
                    cell.border = border
        
            continue
        
        # 读取原表格内容
        单位名称 = row[1].value
        姓名 = row[2].value
        身份证号 = row[3].value
        本年应报销 = float(row[4].value)
        本年已报销 = float(row[5].value)
        本年未报销 = round(float(row[6].value), 2)
        年份 = row[8].value

        # 判断案件号是不是社保数据
        if 'BJ-DT' in row[10].value:
            影像名称 = row[10].value
            信封编号 = 影像名称[-3:]
        else:
            影像名称 = '社保数据'
            信封编号 = ''
            
        # 设置自动换行
        ws.append([nmb-空行-1, 单位名称, 姓名, 身份证号, '个人额度已用完', 影像名称, 0, 信封编号, '部分退单'])
        # nmb+4是因为前面这行是第6行了，nmb是2
        ws.row_dimensions[nmb-空行+4].height = 45
        for row in ws[f'A{nmb-空行+4}:I{nmb-空行+4}']:
            for cell in row:
                # 自动换行、加边框、字体居中
                cell.border = border
                cell.alignment = alignment

        if 选项 == '否':
            continue

        # 退单详情表
        # 创建新的sheet页
        ws_name = wb.create_sheet(title=f'{姓名}')

        # # 设置首行首列为10做为边界空隙，设置第二行行高为5
        # ws_name.row_dimensions[1].height = 15
        # ws_name.column_dimensions['A'].width = 3
        # 设置单元格行高和列宽，行高依次为2:40放图片，列宽依次为B:23，C:11，D:12，E:21
        # ws_name.row_dimensions[2].height = 40
        ws_name.row_dimensions[1].height = 23
        ws_name.row_dimensions[2].height = 23
        ws_name.row_dimensions[3].height = 23
        ws_name.row_dimensions[4].height = 100
        ws_name.row_dimensions[5].height = 23
        ws_name.row_dimensions[6].height = 100
        ws_name.row_dimensions[7].height = 30
        ws_name.column_dimensions['A'].width = 35
        ws_name.column_dimensions['B'].width = 35

        # # 读取图片并添加到工作表并定位到单元格旁边
        # img1 = Image('北京人寿.png')
        # ws_name.add_image(img1, 'B2')

        # 设置单元格格式
        # 第1行标题框
        ws_name.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        ws_name['A1'] = '退单明细表'
        ws_name['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_name['A1'].font = Font(size=14, bold=True)
        # 第2行个人信息框
        ws_name['A2'] = f'被保险人（员工）姓名：{姓名}'
        ws_name['A2'].alignment = Alignment(horizontal='left', vertical='center')
        ws_name['B2'] = f'证件号：{身份证号}'
        ws_name['B2'].alignment = Alignment(horizontal='left', vertical='center')
        # 第3行退单数框
        ws_name['A3'] =  '是否整案退单：否'
        ws_name['A3'].alignment = Alignment(horizontal='left', vertical='center')
        ws_name['B3'] = '退单票据张数：0'
        ws_name['B3'].alignment = Alignment(horizontal='left', vertical='center')
        # 合并第4行，退单原因框
        ws_name.merge_cells(start_row=4, start_column=1, end_row=4, end_column=2)
        ws_name['A4'] = '退单原因：个人额度已用完'
        ws_name['A4'].alignment = Alignment(horizontal='left', vertical='center')
        # 合并第5行，超限额年度行
        ws_name.merge_cells(start_row=5, start_column=1, end_row=5, end_column=2)
        ws_name['A5'] = f'年度：{年份}'
        ws_name['A5'].alignment = Alignment(horizontal='left', vertical='center')
        # 合并第6行，备注框
        ws_name.merge_cells(start_row=6, start_column=1, end_row=6, end_column=2)
        ws_name['A6'] = f'备注：年度累计应报销{本年应报销}元\n     年度已报销金额{本年已报销}元\n     年度未报销金额{本年未报销}元'
        ws_name['A6'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        # 第7行，退单人框
        ws_name['A7'] = '退单人：ZYN'
        ws_name['A7'].alignment = Alignment(horizontal='left', vertical='center')
        ws_name['B7'] = f'确认日期：{dates}'
        ws_name['B7'].alignment = Alignment(horizontal='left', vertical='center')

        # 加边框
        for row in ws_name['A1:B7']:
            for cell in row:
                cell.border = border            

    wb.save('缓存文件夹/地铁超限额退单表.xlsx')
    下载单个文件('缓存文件夹/地铁超限额退单表.xlsx')

def 地铁赔付过万退单表(表格路径, 选项='是'):
    # 样式
    thin = Side(border_style="thin", color="000000")  #边框样式，颜色
    border = Border(left=thin, right=thin, top=thin, bottom=thin)   #边框的位置
    font = Font(size=14, bold=True, name='微软雅黑',  color="FF0000")   #字体大小，加粗，字体名称，字体名字
    fill = PatternFill(patternType="solid", start_color='FBEFF2')  # 填充
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)   # 字体居中并自动换行

    # 开始读取原表格内容
    wb_old = load_workbook(表格路径)
    ws_old = wb_old.active
    put_text(f'共导入{ws_old.max_row-1}条数据')
    nmb = 0
    空行 = 0
    put_processbar('row', auto_close=True)
    for row in ws_old.rows:
        nmb += 1
        set_processbar('row', nmb / ws_old.max_row, label=nmb)

        # 检查空行
        c = 0
        for cell in row:
            if cell.value is not None:
                c = 1
        if c == 0:
            空行 += 1
            continue
        
        if nmb-空行 == 1:
            # 读取当天日期
            dates = datetime.now().strftime('%Y-%m-%d')

            # 创建工作薄
            wb = Workbook()
            ws = wb.active
            ws.title = '退单汇总'

            # 设置首行首列为10做为边界空隙，设置第二行行高为5
            ws.row_dimensions[1].height = 10
            # ws.column_dimensions['A'].width = 2
            # 设置单元格行高和列宽，行高依次为2:60放图片，3:20，4:15，列宽依次为B:3，C:17，D:7，E:20，F:20，G:18，H:6，I:4，J:8
            ws.row_dimensions[2].height = 40
            ws.row_dimensions[3].height = 20
            ws.row_dimensions[4].height = 15
            ws.column_dimensions['A'].width = 3.5
            ws.column_dimensions['B'].width = 17
            ws.column_dimensions['C'].width = 7
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 18
            ws.column_dimensions['G'].width = 8
            ws.column_dimensions['H'].width = 6
            ws.column_dimensions['I'].width = 10

            # 读取图片并添加到工作表并定位到单元格旁边
            img = Image('北京人寿.png')
            ws.add_image(img, 'B2')

            # 合并第3、4行的单元格
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=9)
            ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=9)

            # 写入内容并设置居中和居左
            # 设置粗体
            # 单位名称 = '北京市地铁运营有限公司'
            单位 = f'理赔退单交接表（北京市地铁运营有限公司）'
            ws.cell(row=3, column=1, value=单位).font = Font(size=14, bold=True)
            # 字体居中并循环加入下边框
            ws['A3'].alignment = alignment
            for row in ws['A3:I3']:
                for cell in row:
                    cell.border = Border(bottom=thin)

            # 获取现在的时间并格式化，左对齐
            ws['A4'] = dates
            ws['A4'].alignment = Alignment(horizontal='left', vertical='center')

            # 加入标题
            ws.append(['序号', '单位名称', '姓名', '证件号码', '退单原因', '影像名称', '退单票据份数', '信封编号', '退单类型'])
            # 填充颜色
            for row in ws['A5:I5']:
                for cell in row:
                    # 填充颜色、自动换行、加边框、字体居中
                    cell.fill = fill
                    cell.alignment = alignment
                    cell.border = border
        
            nmb += 1
            continue
        
        # 读取原表格内容
        单位名称 = row[0].value
        姓名 = row[1].value
        身份证号 = row[2].value

        # 判断案件号是不是社保数据
        if 'BJ-DT' in row[5].value:
            影像名称 = row[5].value
            信封编号 = 影像名称[-3:]
        else:
            影像名称 = '社保数据'
            信封编号 = ''
                    
        # 设置自动换行
        ws.append([nmb-空行-1, 单位名称, 姓名, 身份证号, '赔付过万,请提供身份证', 影像名称, 0, 信封编号, '部分退单'])
        # nmb+4是因为前面这行是第6行了，nmb是2
        ws.row_dimensions[nmb-空行+4].height = 45
        for row in ws[f'A{nmb-空行+4}:I{nmb-空行+4}']:
            for cell in row:
                # 自动换行、加边框、字体居中
                cell.border = border
                cell.alignment = alignment

        if 选项 == '否':
            continue

        # 退单详情表
        # 创建新的sheet页
        ws_name = wb.create_sheet(title=f'{姓名}')

        # 设置首行首列为10做为边界空隙，设置第二行行高为5
        ws_name.row_dimensions[1].height = 15
        ws_name.column_dimensions['A'].width = 3
        # 设置单元格行高和列宽，行高依次为2:40放图片，列宽依次为B:23，C:11，D:12，E:21
        ws_name.row_dimensions[2].height = 40
        ws_name.row_dimensions[3].height = 23
        ws_name.row_dimensions[4].height = 23
        ws_name.row_dimensions[5].height = 23
        ws_name.row_dimensions[6].height = 100
        ws_name.row_dimensions[7].height = 100
        ws_name.row_dimensions[8].height = 30
        ws_name.column_dimensions['B'].width = 35
        ws_name.column_dimensions['C'].width = 35

        # 读取图片并添加到工作表并定位到单元格旁边
        img1 = Image('北京人寿.png')
        ws_name.add_image(img1, 'B2')

        # 设置单元格格式
        # 第3行标题框
        ws_name.merge_cells(start_row=3, start_column=2, end_row=3, end_column=3)
        ws_name['B3'] = '退单明细表'
        ws_name['B3'].alignment = Alignment(horizontal='center', vertical='center')
        ws_name['B3'].font = Font(size=14, bold=True)
        # 第4行个人信息框
        ws_name['B4'] = f'被保险人（员工）姓名：{姓名}'
        ws_name['B4'].alignment = Alignment(horizontal='left', vertical='center')
        ws_name['C4'] = f'证件号：{身份证号}'
        ws_name['C4'].alignment = Alignment(horizontal='left', vertical='center')
        # 第5行退单数框
        ws_name['B5'] =  '是否整案退单：否'
        ws_name['B5'].alignment = Alignment(horizontal='left', vertical='center')
        ws_name['C5'] = '退单票据张数：0'
        ws_name['C5'].alignment = Alignment(horizontal='left', vertical='center')
        # 合并第6行，退单原因框
        ws_name.merge_cells(start_row=6, start_column=2, end_row=6, end_column=3)
        ws_name['B6'] = '退单原因：赔付过万，请提供身份证'
        ws_name['B6'].alignment = Alignment(horizontal='left', vertical='center')
        # 合并第7行，备注框
        ws_name.merge_cells(start_row=7, start_column=2, end_row=7, end_column=3)
        ws_name['B7'] = f'备注：'
        ws_name['B7'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        # 第8行，退单人框
        ws_name['B8'] = '退单人：ZYN'
        ws_name['B8'].alignment = Alignment(horizontal='left', vertical='center')
        ws_name['C8'] = f'确认日期：{dates}'
        ws_name['C8'].alignment = Alignment(horizontal='left', vertical='center')

        # 加边框
        for row in ws_name['B4:C8']:
            for cell in row:
                cell.border = border
        
        nmb += 1

    wb.save('缓存文件夹/地铁赔付过万退单表.xlsx')
    下载单个文件('缓存文件夹/地铁赔付过万退单表.xlsx')

# 人寿理算函数
def 获取赔付方案比例(保单详情, 保单号, 单位, 票据日期):
    # 保单详情类型是字典
    # 变量初始化
    赔付比例 = 方案 = ''
    # 获取方案
    if 单位 == '地铁':
        单位赔付比例 = BJRS.赔付比例().get(单位)
    elif 单位 == '公交':
        单位赔付比例 = BJRS.赔付比例().get(单位).get(保单号)
    # 获取赔付比例
    for 方案名称 in 保单详情:
        责任始期 = 保单详情[方案名称]['责任始期']
        责任止期 = 保单详情[方案名称]['责任止期']
        # 票据日期在正常保期内
        if 责任始期 <= 票据日期 <= 责任止期:
            赔付比例 = 单位赔付比例[方案名称]

        # 判断地铁20年6月前票据
        elif 责任始期 > '2020-05-31' and 票据日期 < '2020-06-01':
            赔付比例 = 单位赔付比例[方案名称]

        # 地铁需要方案判断起付线，公交不需要方案
        if 赔付比例:
            if '退休' in 方案名称:
                方案 = '退休'
            else:
                方案 = '在职'
            break

    return 赔付比例, 方案
    
def 单位简称获取(批次号):
    人寿单位名称编码 = BJRS.单位名称()

    try:
        单位编码 = ''
        for i in 批次号[3:8]:
            if is_number(i) == False:
                单位编码 += i
        单位简称 = 人寿单位名称编码[单位编码][3]
        单位保单号 = ''
        单位保单号简称 = ''
        for i in 人寿单位名称编码[单位编码][1]:
            单位保单号 += i
            单位保单号简称 += f'{i[-4:]},'
    except:
        单位简称 = 单位保单号简称 = 单位保单号 = ''   
        
    return 单位简称, 单位保单号简称, 单位保单号

def 地铁理算公式(自付一, 起付金额, 自付二, 赔付比例, 票据类别, 年份, 方案, 补贴='', 累计自付一=0, 累计自付二=0, 累计起付金额=0):
    # 22年票据不能个案理算，只能历史理算
    if 票据类别 == '门诊':
        if 方案 == '在职':
            起付线 = 1800
            补贴额 = 800
        else:
            起付线 = 1300
            补贴额 = 650
        
        # 21年以后理算
        if 年份 > '21':
            # 判断首次补贴
            if 补贴 == 0:
                # 达到赔付标准理算
                if 累计自付一 + 自付一 >= 起付线:
                    补贴 = 1
                    理算 = (累计自付一 + 自付一 + 自付二 - 起付线 + 补贴额) * 赔付比例 + 累计自付二
                    累计自付一 = 累计自付二 = 0
                # 不到赔付标准累计
                else:
                    累计自付一 += 自付一
                    累计自付二 += 自付二 * 赔付比例
                    理算 = 0
                
            elif 补贴 == 1:
                # 补贴过后直接理算
                理算 = (累计自付一 + 自付一 + 自付二) * 赔付比例 + 累计自付二

            # 个案理算
            else:
                # 判断一下补贴（不一定准）
                if 自付一 > 起付金额 > 0: 
                    理算 = (自付一 + 自付二 - 起付金额 + 补贴额) * 赔付比例
                #　如果相等并大于0，累计
                elif 自付一 <= 起付金额 > 0:
                    理算 = 0
                # 如果其它，理算
                else:
                    理算 = (自付一 + 自付二 - 起付金额) * 赔付比例
                

        # 22年以前公式
        else:
            # 判断首次补贴
            if 补贴 == 0:
                # 判断自付一是否大于起付额且大于等于0
                if 累计自付一 + 自付一 > 起付金额 + 累计起付金额 >= 0:
                    补贴 = 1
                    理算 = ((累计自付一 + 自付一) + 自付二 - (起付金额 + 累计起付金额) + 补贴额) * 赔付比例 + 累计自付二
                    累计自付一 = 累计自付二 = 累计起付金额 = 0
                # 不达到赔付标准累计
                else:
                    累计自付一 += 自付一
                    累计自付二 += 自付二 * 赔付比例
                    累计起付金额 += 起付金额
                    理算 = 0
            # 有补贴
            elif 补贴 == 1:
                # 补贴后直接理算
                理算 = (自付一 + 自付二 - 起付金额) * 赔付比例

            # 个案理算
            else:
                # 判断一下补贴（不一定准）
                if 自付一 > 起付金额 > 0: 
                    理算 = (自付一 + 自付二 - 起付金额 + 补贴额) * 赔付比例
                #　如果相等并大于0，累计
                elif 自付一 <= 起付金额 > 0:
                    理算 = 0
                # 如果其它，理算
                else:
                    理算 = (自付一 + 自付二 - 起付金额) * 赔付比例
                
    # 除门诊外，住院和门特都一样
    else:
        理算 = (自付一 + 自付二) * 赔付比例

    return 理算, 补贴, 累计自付一, 累计自付二, 累计起付金额
    
def 公交理算公式(自付一, 起付金额, 超封顶金额, 自付二, 自费, 赔付比例, 票据类别):
    ''' 
        0门诊自付一、1门诊自付二、2门诊自费、3门诊超封顶金额5万以内、4门诊超封顶金额10万以内、5门诊超封顶金额15万以内、6门诊超封顶金额25万以内
        7住首自付一、8住首自付二 、9住首自费、10住非自付一 、11住非自付二、12住非自费
        13住院超封顶金额5万以内、14住院超封顶金额10万以内、15住院超封顶金额15万以内、16住院超封顶金额25万以内
        17门特自付一、18门特自付二、19门特自费
    '''

    if 票据类别 == '门诊':
        # 先判断是不是家属
        if len(赔付比例) > 5:
            # 门诊超封顶金额25万以内
            if 250000 >= 超封顶金额 > 150000:
                超封顶赔付比例 = 赔付比例[6]
            # 门诊超封顶金额15万以内
            elif 超封顶金额 > 100000:
                超封顶赔付比例 = 赔付比例[5]
            # 门诊超封顶金额10万以内
            elif 超封顶金额 > 50000:
                超封顶赔付比例 = 赔付比例[4]
            # 门诊超封顶金额5万以内
            else:
                超封顶赔付比例 = 赔付比例[3]
            # 理算公式
            理算 = (自付一-起付金额)*赔付比例[0] +\
                自付二*赔付比例[1] + 自费*赔付比例[2] +\
                    超封顶金额*超封顶赔付比例
        else:
            # 家属只赔付自付一，0起付线下，1起付线上
            理算 = 起付金额*赔付比例[0] + (自付一-起付金额+自付二)*赔付比例[1]

    elif 票据类别 == '住院':
        # 先判断是不是家属
        if len(赔付比例) > 5:
            # 住院超封顶金额25万以内
            if 250000 >= 超封顶金额 > 150000:
                超封顶赔付比例 = 赔付比例[16]
            # 住院超封顶金额15万以内
            elif 超封顶金额 > 100000:
                超封顶赔付比例 = 赔付比例[15]
            # 住院超封顶金额10万以内
            elif 超封顶金额 > 50000:
                超封顶赔付比例 = 赔付比例[14]
            # 住院超封顶金额5万以内
            else:
                超封顶赔付比例 = 赔付比例[13]
            # 是否是首次住院
            if 起付金额 == 1300 or 起付金额 == -1300:
                理算 = (自付一-起付金额)*赔付比例[7] +\
                    自付二*赔付比例[8] + 自费*赔付比例[9] + 超封顶金额*超封顶赔付比例
            else:
                理算 = (自付一-起付金额)*赔付比例[10] +\
                    自付二*赔付比例[11] + 自费*赔付比例[12] + 超封顶金额*超封顶赔付比例
        else:
            # 家属只赔付自付一，0起付线下，1起付线上
            理算 = 起付金额*赔付比例[0] + (自付一-起付金额+自付二)*赔付比例[1]
        
    elif 票据类别 == '门特':
        # 先判断是不是家属
        if len(赔付比例) > 5:
            理算 = (自付一-起付金额) + 自付二 + 自费*0.4
        else:
            理算 = 0
        
    return 理算

def 北京人寿地铁个案理算(案件号, 年度基础数据='', 查询年份='', 获取系统理算='是'):
    # 获取案件信息
    url = GD.案件号查询网址(案件号)
    data1 = GD.获取案件信息(url, headers)

    if data1 == '没有更多啦~':
        put_text(f'系统没有查询到此案件号{案件号}')
        return data1

    x = data1['data'][0]
    # 标识用来记录有没有获取赔付比例，0是没有（需要获取），1是有（跳过）;跳过用来判断是否继续进行，0是要跳过，1是继续；不同年份记录公交的不同年份
    门诊理算 = 住院理算 = 门特理算 = 0
    票据年份 = 门诊原赔付比例 = 门诊赔付比例 = 住院原赔付比例 = 住院赔付比例 = 门特原赔付比例 = 门特赔付比例 = ''

    个案列表 = GD.提取案件列表个案详情(x)
    # 有指定的查询年份并且上传年份小于查询年份直接略过
    if 查询年份 and 查询年份 > 个案列表['上传时间'][2:4]:
        return '', 0, '', 年度基础数据

    # 标题：姓名，身份证号，案件号，上传时间，回传时间，门诊理算，住院理算，门特理算，理算合计，系统理算，年份，历史自付二
    lst1 = [个案列表['姓名'], 个案列表['身份证号'], 个案列表['案件号'], 个案列表['上传时间'][:10], 个案列表['回传时间'][:10]]

    # 获取基础信息
    url = GD.案件详情查询网址(个案列表['案件id'])
    个案信息 = GD.获取案件信息(url, headers)   # 获取个案详情
    if 个案信息:
        基础信息 = GD.提取案件详情基础信息(个案信息)
    else:
        lst1.append('案件错误')
        案件号 = 个案列表['案件号']
        put_text(f'{案件号}错误！！！！！')
        return lst1, 0, '', 年度基础数据

    # 判断问题件如果是问题件，就返回空
    if 基础信息['问题件'] == '是':
        lst1.append('问题件')
        return lst1, 0, '', 年度基础数据
        
    if 获取系统理算 == '否':
        # 没理算案件跳过
        if 个案列表['理算状态'] != '已理算':
            lst1.append('未理算')
            return lst1, 0, '', 年度基础数据

    # 获取保单方案和保期
    保单号 = 基础信息.get('保单号')
    # 没有保单就跳过
    if not 保单号:
        return '', 0, '', 年度基础数据
    保期查询号 = 基础信息.get('保期查询号')
    url = GD.案件详情保期查询网址(保期查询号, 保单号)
    保单信息 = GD.获取案件信息(url, headers)
    保单详情 = GD.提取保期信息(保单信息)
    # 未获取到保单详情就跳过
    if not 保单详情:
        return '', 0, '', 年度基础数据

    # 检测票据明细内容，单张票据理算
    # 是否有门诊
    if 个案信息['stub']:
        票据类别 = '门诊'
        方案变更 = ''
        # 初始化各种数据
        标识 = 自付一 = 起付金额 = 自付二 = 0
        # 获取合计信息，判断有没有票据
        票据信息 = GD.提取案件详情城镇门诊合计信息(个案信息['sum_stub'])
        if 票据信息['合计'] == 0:
            合计标识 = 0
        else:
            合计标识 = 1

        for 门诊信息 in 个案信息['stub']:
            票据信息 = GD.提取案件详情城镇门诊信息(门诊信息)
            if 基础信息.get('医保类型') == '城镇居民':
                票据信息['票据类型'] = 1
            # 如果退单跳过，为防止都退单没有标问题件导致获取赔付比例失败的情况，暂时标记跳过为1，获取后改为0
            if 票据信息['退单状态'] == 1:
                continue

            票据时间 = 票据信息['票据时间']

            # 小于20年的票据跳过
            if 票据时间[2:4] < '20':
                continue
                
            票据年份 = 票据时间[2:4]
            # 票据合计为0时只取票据年份
            if 合计标识 == 0:
                break
            # 如果有明确的查询年份且不在查询的年份里，跳过
            if 查询年份 and 查询年份 != 票据年份:
                continue

            # 获取赔付比例和方案
            门诊赔付比例, 方案 = 获取赔付方案比例(保单详情, 保单号, '地铁', 票据时间)
            # 有赔付比例不跳过
            if not 门诊赔付比例:
                # print(f'{案件号}门诊票据不在保期内！！！！！')
                continue

            # 记录赔付比例是否有变更，有变更就是方案有变化
            if 标识 == 0:
                门诊原赔付比例 = 门诊赔付比例
                标识 = 1
            
            # 方案变更，分别累计
            if 门诊原赔付比例 == 门诊赔付比例:
                自付一 += 票据信息['自付一']
                起付金额 += 票据信息['起付金额']
                自付二 += 票据信息['自付二']
            else:
                if not 方案变更:
                    变更前自付一 = 自付一
                    变更前起付金额 = 起付金额
                    变更前自付二 = 自付二
                    自付一 = 票据信息['自付一']
                    起付金额 = 票据信息['起付金额']
                    自付二 = 票据信息['自付二']
                    方案变更 = '是'
                else:
                    自付一 += 票据信息['自付一']
                    起付金额 += 票据信息['起付金额']
                    自付二 += 票据信息['自付二']

        if 合计标识 == 1:
            # 没有票据年份证明唯一的票据退单了
            if 票据年份:
                # 判断是否是个案
                if 年度基础数据 == '':
                    年度基础数据[票据年份] = ['', 0, 0, 0, 0]

                elif 年度基础数据.get(票据年份) == None:
                    年度基础数据[票据年份] = [0, 0, 0, 0, 0]

                if 门诊赔付比例:
                    if not 方案变更:
                        # 自付一, 起付金额, 自付二, 赔付比例, 票据类别, 年份, 方案, 补贴, 累计自付一, 累计自付二, 累计起付金额'
                        门诊理算, 补贴, 累计自付一, 累计自付二, 累计起付金额 = 地铁理算公式(自付一, 起付金额, \
                            自付二, 门诊赔付比例, 票据类别, 票据年份, 方案, 年度基础数据[票据年份][0], \
                                年度基础数据[票据年份][1], 年度基础数据[票据年份][2], 年度基础数据[票据年份][3])
                        # 写入临时数据
                        年度基础数据[票据年份] = [补贴, 累计自付一, 累计自付二, 累计起付金额]
                    else:
                        # 理算变更前的结果
                        变更前门诊理算, 补贴, 累计自付一, 累计自付二, 累计起付金额 = 地铁理算公式(变更前自付一, 变更前起付金额, \
                            变更前自付二, 门诊原赔付比例, 票据类别, 票据年份, 方案, 年度基础数据[票据年份][0], \
                                年度基础数据[票据年份][1], 年度基础数据[票据年份][2], 年度基础数据[票据年份][3])
                        # 写入临时数据
                        年度基础数据[票据年份] = [补贴, 累计自付一, 累计自付二, 累计起付金额]
                        # 理算变更后的结果
                        门诊理算, 补贴, 累计自付一, 累计自付二, 累计起付金额 = 地铁理算公式(自付一, 起付金额, \
                            自付二, 门诊赔付比例, 票据类别, 票据年份, 方案, 年度基础数据[票据年份][0], \
                                年度基础数据[票据年份][1], 年度基础数据[票据年份][2], 年度基础数据[票据年份][3])
                        # 写入临时数据
                        年度基础数据[票据年份] = [补贴, 累计自付一, 累计自付二, 累计起付金额]

                        # 两次结果相加
                        门诊理算 += 变更前门诊理算

                    lst1.append(round(门诊理算, 2))
                else:
                    lst1.append('无赔付比例')
        else:
            lst1.append(0)
                    
    else:
        lst1.append('-')

    # 是否有住院
    if 个案信息['stub_hospital']:
        票据类别 = '住院'
        方案变更 = ''
        标识 = 自付一 = 起付金额 = 自付二 = 0
        # 获取合计信息，判断有没有票据
        票据信息 = GD.提取案件详情住院合计信息(个案信息['sum_stub_hospital'])
        if 票据信息['合计'] == 0:
            合计标识 = 0
        else:
            合计标识 = 1

        for 住院信息 in 个案信息['stub_hospital']:
            票据信息 = GD.提取案件详情住院信息(住院信息)
            # 如果退单跳过，为防止都退单没有标问题件导致获取赔付比例失败的情况，暂时标记跳过为1，获取后改为0
            if 票据信息['退单状态'] == 1:
                continue
            # 获取票据时间和年份
            票据时间 = 票据信息['出院时间']
            # 小于20年的票据跳过
            if 票据时间[2:4] < '20':
                continue
    
            票据年份 = 票据时间[2:4]
            # 票据合计为0时只取票据年份
            if 合计标识 == 0:
                break          

            # 如果有明确的查询年份且不在查询的年份里，跳过
            if 查询年份 and 查询年份 != 票据年份:
                continue

            # 获取赔付比例和方案
            住院赔付比例, 方案 = 获取赔付方案比例(保单详情, 保单号, '地铁', 票据时间)
            # 有赔付比例不跳过
            if not 住院赔付比例:
                # print(f'{案件号}住院票据不在保期内！！！！！')
                continue
            # 记录赔付比例是否有变更，有变更就是方案有变化
            if 标识 == 0:
                住院原赔付比例 = 住院赔付比例
                标识 = 1

            # 方案变更，分别累计
            if 住院原赔付比例 == 住院赔付比例:
                自付一 += 票据信息['自付一']
                起付金额 += 票据信息['起付金额']
                自付二 += 票据信息['自付二']
            else:
                if not 方案变更:
                    变更前自付一 = 自付一
                    变更前起付金额 = 起付金额
                    变更前自付二 = 自付二
                    自付一 = 票据信息['自付一']
                    起付金额 = 票据信息['起付金额']
                    自付二 = 票据信息['自付二']
                    方案变更 = '是'
                else:
                    自付一 += 票据信息['自付一']
                    起付金额 += 票据信息['起付金额']
                    自付二 += 票据信息['自付二']

        if 合计标识 == 1:
            # 没有票据年份证明唯一的票据退单了
            if 票据年份:
                if 住院赔付比例:
                    if not 方案变更:
                        住院理算, 补贴, 累计自付一, 累计自付二, 累计起付金额 = 地铁理算公式(自付一, 起付金额, \
                            自付二, 住院赔付比例, 票据类别, 票据年份, 方案)
                    else:
                        变更前住院理算, 补贴, 累计自付一, 累计自付二, 累计起付金额 = 地铁理算公式(变更前自付一, 变更前起付金额, \
                            变更前自付二, 住院原赔付比例, 票据类别, 票据年份, 方案)
                        
                        住院理算, 补贴, 累计自付一, 累计自付二, 累计起付金额 = 地铁理算公式(自付一, 起付金额, \
                            自付二, 住院赔付比例, 票据类别, 票据年份, 方案)

                        住院理算 += 变更前住院理算

                    lst1.append(round(住院理算, 2))
                else:
                    if '无赔付比例' not in lst1:
                        lst1.append('无赔付比例')
                        
        else:
            lst1.append(0)
    else:
        lst1.append('-')
            
    # 是否有门特
    if 个案信息['stub_whole']:
        票据类别 = '门特'
        方案变更 = ''
        标识 = 自付一 = 起付金额 = 自付二 = 0
        # 获取合计信息，判断有没有票据
        票据信息 = GD.提取案件详情门特合计信息(个案信息['sum_stub_whole'])
        if 票据信息['合计'] == 0:
            合计标识 = 0
        else:
            合计标识 = 1
            
        for 门特信息 in 个案信息['stub_whole']:
            票据信息 = GD.提取案件详情门特信息(门特信息)
            # 如果退单跳过，为防止都退单没有标问题件导致获取赔付比例失败的情况，暂时标记跳过为1，获取后改为0
            if 票据信息['退单状态'] == 1:
                continue

            票据时间 = 票据信息['票据时间']
            
            # 小于20年的票据跳过
            if 票据时间[2:4] < '20':
                continue
                
            票据年份 = 票据时间[2:4]
            # 票据合计为0时只取票据年份
            if 合计标识 == 0:
                break          

            # 如果有明确的查询年份且不在查询的年份里，跳过
            if 查询年份 and 查询年份 != 票据年份:
                continue

            # 获取赔付比例和方案
            门特赔付比例, 方案 = 获取赔付方案比例(保单详情, 保单号, '地铁', 票据时间)
            # 有赔付比例不跳过
            if not 门特赔付比例:
                # print(f'{案件号}门特票据不在保期内！！！！！')
                continue
            # 记录赔付比例是否有变更，有变更就是方案有变化
            if 标识 == 0:
                门特原赔付比例 = 门特赔付比例
                标识 = 1
            # 方案变更，分别累计
            if 门特原赔付比例 == 门特赔付比例:
                自付一 += 票据信息['自付一']
                起付金额 += 票据信息['起付金额']
                自付二 += 票据信息['自付二']
            else:
                if not 方案变更:
                    变更前自付一 = 自付一
                    变更前起付金额 = 起付金额
                    变更前自付二 = 自付二
                    自付一 = 票据信息['自付一']
                    起付金额 = 票据信息['起付金额']
                    自付二 = 票据信息['自付二']
                    方案变更 = '是'
                else:
                    自付一 += 票据信息['自付一']
                    起付金额 += 票据信息['起付金额']
                    自付二 += 票据信息['自付二']

        
        if 合计标识 == 1:
            # 没有票据年份证明唯一的票据退单了
            if 票据年份:
                if 门特赔付比例:
                    if not 方案变更:
                        门特理算, 补贴, 累计自付一, 累计自付二, 累计起付金额 = 地铁理算公式(自付一, 起付金额, \
                            自付二, 门特赔付比例, 票据类别, 票据年份, 方案)
                    else:
                        变更前门特理算, 补贴, 累计自付一, 累计自付二, 累计起付金额 = 地铁理算公式(变更前自付一, 变更前起付金额, \
                            变更前自付二, 门特原赔付比例, 票据类别, 票据年份, 方案)
                        
                        门特理算, 补贴, 累计自付一, 累计自付二, 累计起付金额 = 地铁理算公式(自付一, 起付金额, \
                            自付二, 门特赔付比例, 票据类别, 票据年份, 方案)

                        门特理算 += 变更前门特理算
                    
                    lst1.append(round(门特理算, 2))
                else:
                    if '无赔付比例' not in lst1:
                        lst1.append('无赔付比例')
        else:
            lst1.append(0)

    else:
        lst1.append('-')
    
    理算合计 = 门诊理算 + 住院理算 + 门特理算
    lst1.append(round(理算合计, 2))

    # 是否获取系统理算数据
    if 获取系统理算 == '是':
        if 个案列表['理算状态'] == '已理算':
            try:
                # 获取理算结果
                url = GD.列表页理算结果查询网址(个案列表['案件号'])
                理算信息 = GD.获取案件信息(url, headers)
                if 理算信息 == '没有理算的责任结果':
                    lst1.append('-')
                    lst1.append('-')
                    lst1.append('-')
                    lst1.append('-')
                else:
                    理算结果 = GD.提取列案件表页理算结果(理算信息)
                    lst1.append(理算结果['门诊回传金额'])
                    lst1.append(理算结果['住院回传金额'])
                    lst1.append(理算结果['门特回传金额'])
                    lst1.append(理算结果['回传总额'])
            
            except:
                lst1.append('-')
                lst1.append('-')
                lst1.append('-')
                lst1.append('失败')
        else:
            lst1.append('-')
            lst1.append('-')
            lst1.append('-')
            lst1.append('未理算')
            
    lst1.append(票据年份)

    # 累计自付二理算结果，判断有没有门诊
    if 年度基础数据.get(票据年份):
        lst1.append(年度基础数据[票据年份][2])
    else:
        lst1.append(0)

    try:
        # 如果赔付比例不一样，添加方案变更字样
        if 门诊原赔付比例 != 门诊赔付比例 or 住院原赔付比例 != 住院赔付比例 or 门特原赔付比例 != 门特赔付比例:
            lst1.append('是')
        else:
            lst1.append('-')
    except:
        lst1.append('失败')

    return lst1, 理算合计, 票据年份, 年度基础数据

def 北京人寿个人历史理算查询(身份证号, 单位, 查询年份, 获取系统理算='是'):
    lst = []
    本年应赔 = 0
    # 获取案件信息
    url = GD.身份证号查询网址(身份证号)
    data1 = GD.获取案件信息(url, headers)

    if data1 == '没有更多啦~':
        lst.append(['', 身份证号, '系统没有此人'])
        return lst

    # 案件总数 = data1['page']['count']
    案件总页数 = data1['page']['pages']
    data2 = data1['data']
    if 案件总页数 > 1:
        # 按页循环获取数据
        for page in range(2, 案件总页数+1):
            url = GD.身份证号查询网址(身份证号, 10, page)
            data1 = GD.获取案件信息(url, headers)   # 获取案件列表页
            # 把每页的数据合并到一起
            for i in data1['data']:
                data2.append(i)

    if data2:
        # 22年需要累计自付一和自付二，22年之前的只累计自付二
        # 如果有条件查询，去掉最后的年字
        if 查询年份:
            查询年份 = 查询年份[:2]
        # 如果全部查询，要分年度累计赔付金额
        else:
            年度应赔 = {'20': 0, '21': 0, '22': 0}

        if 单位 == '地铁':
            # 记录年度补贴：格式{[补贴，累计自付一，累计自付二，历史累计自付二，累计自费]}
            年度基础数据 = {}
        nmb = 0
        put_processbar('案件号', auto_close=True)
        # 倒序获取
        for x in data2[::-1]:
            nmb += 1
            set_processbar('案件号', nmb / len(data2), label=nmb)
            time.sleep(0.1)
            个案列表 = GD.提取案件列表个案详情(x)
            if 单位 == '地铁':
                lst1, 理算合计, 票据年份, 年度基础数据 = 北京人寿地铁个案理算(个案列表['案件号'], 年度基础数据, 查询年份, 获取系统理算)

                # 判断lst1是否为空，空就是跳过
                if not lst1:
                    continue
                # 如果未理算，提示未理算
                if '未理算' in lst1 or '问题件' in lst1:
                    lst.append(lst1)
                    continue
                # 本年度应报销
                本年应赔 += 理算合计
                if 查询年份:
                    lst1.append(round(本年应赔, 2))
                elif 票据年份:
                    年度应赔[票据年份] += 理算合计
                    lst1.append(round(年度应赔[票据年份], 2))
                else:
                    lst1.append('-')

                # 查询本年度赔付总额
                try:
                    if 赔付明细.get(身份证号).get(f'20{票据年份}'):
                        lst1.append(round(赔付明细.get(身份证号).get(f'20{票据年份}').get('年度已报销'), 2))
                    else:
                        lst1.append(0)
                except:
                    lst1.append('-')
                    
                lst.append(lst1)

            # 如果列表还没有要查询的年份案件，提示无
            if not lst:
                lst.append([个案列表['姓名'], 身份证号, f'无{查询年份}年票据'])

    else:
        lst.append(['', 身份证号, '导出个人案件错误'])

    return lst

def 北京人寿批次理算(批次号, 单位):
    lst = []
    # 获取案件信息
    url = GD.批次号查询网址(批次号)
    data1 = GD.获取案件信息(url, headers)

    if data1 == '没有更多啦~':
        put_text(f'系统没有查询到此批次号{批次号}')
        return data1

    案件总数 = data1['page']['count']
    案件总页数 = data1['page']['pages']
    put_text(f'批次{批次号}共{案件总数}件案件。\n')
    data2 = data1['data']
    if 案件总页数 > 1:
        # 按页循环获取数据
        for page in range(2, 案件总页数+1):
            url = GD.批次号查询网址(批次号, 10, page)
            data1 = GD.获取案件信息(url, headers)   # 获取案件列表页
            # 把每页的数据合并到一起
            for i in data1['data']:
                data2.append(i)

    n = 0
    put_processbar(批次号, auto_close=True)
    for x in data2:
        time.sleep(0.1)
        n += 1
        set_processbar(批次号, n / 案件总数, label=n)
        个案列表 = GD.提取案件列表个案详情(x)
        lst1, 理算合计, 票据年份, 年度基础数据 = 北京人寿地铁个案理算(个案列表['案件号'])

        lst.append(lst1)

    return lst

def 北京人寿公交问题件理算(案件号, 方案编码):
    # 获取案件信息
    url = GD.案件号查询网址(案件号)
    data1 = GD.获取案件信息(url, headers)

    if data1 == '没有更多啦~':
        put_text(f'系统没有查询到此案件号{案件号}')
        lst1 = [[f'系统没有查询到此案件号{案件号}']]
        return lst1

    x = data1['data'][0]
    门诊理算 = 住院理算 = 门特理算 = 0

    个案列表 = GD.提取案件列表个案详情(x)

    # 标题：姓名，身份证号，案件号，上传时间，回传时间，门诊理算，住院理算，门特理算，理算合计，系统理算，年份，历史自付二
    lst1 = [个案列表['姓名'], 个案列表['身份证号'], 个案列表['案件号'], 个案列表['上传时间'][:10], 个案列表['回传时间'][:10]]

    # 获取基础信息
    url = GD.案件详情查询网址(个案列表['案件id'])
    个案信息 = GD.获取案件信息(url, headers)   # 获取个案详情

    # 获取赔付比例
    编码对应方案 = {'G1002': '参加市总工会',
                    'G1003': '未参加市总工会',
                    'G1004': '退休人员',
                    'G1005': '建国前参加革命工作',
                    'G1006': '退职人员',
                    'G1007': '最低生活保障标准的在职员工',
                    'G1008': '最低生活保障标准的退休人员',
                    'G1009': '精神病患者（在职，参加工会）',
                    'G1010': '精神病患者（退休）',
                    'G1011': '工伤职业病人员（退休）',
                    'G1012': '视同工伤职业病人员（退休）',
                    'G1013': '有社保的家属（职工的父母）',
                    'G1014': '有社保的家属（双职工的子女）',
                    'G1015': '有社保的家属（单职工的子女）',
                    'G1016': '有社保的家属（单亲职工子女）',
                    'G1017': '无社保的家属（职工的父母）',
                    'G1018': '无社保的家属（双职工的子女）',
                    'G1019': '无社保的家属（单职工的子女）',
                    'G1020': '无社保的家属（单亲职工子女）'}
    方案 = 编码对应方案.get(方案编码)
    各保单赔付比例 = BJRS.赔付比例().get('公交')
    for i in 各保单赔付比例:
        if '6856' in i:
            赔付比例 = 各保单赔付比例[i][方案]

    # 检测票据明细内容，单张票据理算
    # 是否有门诊
    if 个案信息['stub']:
        票据类别 = '门诊'
        # 获取合计信息，判断有没有票据
        票据信息 = GD.提取案件详情城镇门诊合计信息(个案信息['sum_stub'])
        自付一, 起付金额, 超封顶金额, 自付二, 自费 = 票据信息['自付一'], 票据信息['起付金额'], 票据信息['超封顶金额'], 票据信息['自付二'], 票据信息['自费']
        门诊理算 = 公交理算公式(自付一, 起付金额, 超封顶金额, 自付二, 自费, 赔付比例, 票据类别)
        lst1.append(round(门诊理算, 2))

    else:
        lst1.append('-')

    # 是否有住院
    if 个案信息['stub_hospital']:
        票据类别 = '住院'
        
        for 住院信息 in 个案信息['stub_hospital']:
            票据信息 = GD.提取案件详情住院信息(住院信息)
            if 票据信息['退单状态'] == 1:
                continue
            自付一, 起付金额, 超封顶金额, 自付二, 自费 = 票据信息['自付一'], 票据信息['起付金额'], 票据信息['超封顶金额'], 票据信息['自付二'], 票据信息['自费']
            住院理算1 = 公交理算公式(自付一, 起付金额, 超封顶金额, 自付二, 自费, 赔付比例, 票据类别)
            住院理算 += 住院理算1
        lst1.append(round(住院理算, 2))
    else:
        lst1.append('-')
            
    # 是否有门特
    if 个案信息['stub_whole']:
        票据类别 = '门特'
        # 获取合计信息，判断有没有票据
        票据信息 = GD.提取案件详情门特合计信息(个案信息['sum_stub_whole'])
        自付一, 起付金额, 超封顶金额, 自付二, 自费 = 票据信息['自付一'], 票据信息['起付金额'], 票据信息['超封顶金额'], 票据信息['自付二'], 票据信息['自费']
        门特理算 = 公交理算公式(自付一, 起付金额, 超封顶金额, 自付二, 自费, 赔付比例, 票据类别)
        lst1.append(round(门特理算, 2))

    else:
        lst1.append('-')
    
    理算合计 = 门诊理算 + 住院理算 + 门特理算
    lst1.append(round(理算合计, 2))
    
    return lst1


# 以下是页面操作
def 选择操作项目(列表):
    answer = radio("请选择需要操作的项目", options=列表)
    # put_text(f'你选择的项目是：{answer}')
    return answer

def 按钮选择(提示, 列表):
    confirm = actions(提示, 列表, help_text='直接点击按钮选择')
    # put_markdown(f'你选择的是：{confirm}')
    return confirm

def 上传单个文件(选择):
    # onefile = file_upload(选择, required=True, placeholder='选择表格文件', accept=[".xlsx", ".xls"])
    onefile = file_upload(选择, placeholder='选择表格文件', accept=[".xlsx", ".xls"])
    file_name = onefile['filename']
    文件路径 = f'缓存文件夹/{file_name}'
    open(文件路径, 'wb').write(onefile['content'])
    return 文件路径

def 上传多个文件(选择):
    onefile = file_upload(选择, required=True, multiple=True, placeholder='选择TXT文件', accept=[".txt", ".TXT"])
    # print(onefile)
    文件路径列表 = []
    for i in onefile:
        file_name = i['filename']
        文件路径 = f'缓存文件夹/{file_name}'
        # print(文件路径)
        open(文件路径, 'wb').write(i['content'])
        文件路径列表.append(文件路径)

    return 文件路径列表

def 下载单个文件(name):
    if '.xlsx' not in name:
        name = f'{name}.xlsx'
        
    if '缓存文件夹/' in name:
        name = name[6:]

    if os.path.isfile(f'缓存文件夹/{name}'):    #如果path是一个存在的文件，返回True。否则返回False
        if '缓存文件夹' in name:
            content = open(name, 'rb').read()
            put_file(name[6:], content, f'点我下载：{name[6:]}')
        else:
            content = open(f'缓存文件夹/{name}', 'rb').read()
            put_file(name, content, f'点我下载：{name}')
    else:
        put_text(f'没有{name}文件！！！')

def 批量选择项目(name):
    # with put_collapse(f'请上传{name}信息，点击查看模板：'):
    #     put_table([
    #         [f'{name}'],
    #         ['XXXXXXXXXX'],
    #         ['XXXXXXXXXX'],
    #     ])
    
    while True:
        批量列表 = input(f'请输入{name}', required=False, placeholder='输入空是返回上一级')
        if len(批量列表) > 10:
            # 分割成列表
            批量列表 = 批量列表.split()
            # 去重复字符
            l = []
            [l.append(i) for i in 批量列表 if not i in l]
            批量列表 = l
            put_text(f'共输入了{len(批量列表)}件{name}。')
            break

        elif not 批量列表:
            break

        else:
            popup(f'{name}输入位数小于十位，请重新选择输入。')

    return 批量列表

def 是否导入团批():
    # with put_collapse('请上传团批信息，点击查看团批模板：'):
    #     put_table([
    #         ['序号', '姓名', '身份证号', '案件号', '票据数'],
    #         ['001', 'XXX', 'XXXXXXXX', 'XXXXXX', 'XX'],
    #         ['002', 'XXX', 'XXXXXXXX', 'XXXXXX', 'XX'],
    #     ])
    while True:
        answer = 按钮选择('是否导入团批信息？', ['是', '否'])
        # put_text(f'你的选择是：{answer}')
        if answer == '是':
            try:
                onefile = 上传单个文件('请上传团批信息')
                团批信息 = 导入团批信息(onefile)
                break
            except:
                团批信息 = ''
                popup('你没有选择团批文件或者是文件错误')
        else:
            团批信息 = ''
            break
    
    return 团批信息

def 字典转文本(di):
    L = ''  # 临时用
    L1 = ''  # 临时用
    for i in di:
        if di[i]:
            j = '，'.join(di[i])
            L = f'{j}{i};'
            L1 += L
    return L1

def main():
    img = open('zkrj.ico', 'rb').read()
    put_image(img, width='150px', height='80px')
    with put_collapse('欢迎使用中科助手1.6.15版——刷新页面即可回到首页！点开可查看更新详情！'):
        put_text('2022-07-21：1、加入返回上一级功能。')
        put_text('2022-06-15：1、TXT转表格功能优化，自动查询文本编码，可读取大部分文本。')
        put_text('2022-06-09：1、查找历史文件改为模糊查询，只输入关键字即可。')
        put_text('2022-06-08：1、公交社保问题件案件理算增加输入空值检测，直接点提交提示输入内容。')
        put_text('2022-06-06：1、增加公交社保问题件案件理算')
        put_text('2022-05-20：1、输入内容取消上传文件，直接输入文本；2、输入的文本去除多余空格和去重')
        put_text('2022-04-10：1、优化地铁理算流程，输出更详细情况')
        put_text('2022-04-08：1、优化地铁理算流程，把地铁和公交个案理算分开，不混在一起处理；2、修改地铁超限额退单表个人页')
        put_text('2022-04-07：1、优化地铁理算流程，不会在报错导致理算失败')
        put_text('2022-04-01：1、优化获取案件方式；2、优化地铁理算流程；3、修改地铁超限额退单表格式')
        put_text('2022-03-30：1、优化内部结构，把重复个案理算合并到一个使用，修改更方便简洁')
        put_text('2022-03-26：1、案件列表页由一次性获取全部改为循环获取每页数量，减少因超时导致的获取数据失败；2、理算方式由个案合计理算改为个票理算；3、增加每个模块的运行时间。')
        put_text('2022-03-24：优化若干小功能')
        put_text('2022-03-19：1、北京人寿所有单位都已录入，除其它单位没有录入单位简称；2、优化单位信息获取方式，获取理全面；3、优化基础检查保单号的校验，校验理准确')
        put_text('2022-03-18：基础字段校验改为校验保单号（只限6856），保单错误会标红，其它保单显示已关联保单数量')
        put_text('2022-03-16：修复检查筛重文件检查纸质票据号是用案件号判断')
        put_text('2022-03-15：1、增加地铁超限额表转退单表功能；2、优化一些小功能')
        put_text('2022-03-09：1、增加批次号导出基础信息时可选上传时间和理算状态；2、优化一些小功能')
        put_text('2022-03-06：1、增加公交理算和批次理算功能；2、优化一些小功能')
        put_text('2022-03-01：增加身份证号查询个人所有案件或指定日期案件功能，将功能集合到查询基础信息')
        put_text('2022-02-27：优化地铁个案自动理算功能')
        put_text('2022-02-26：优化地铁自动理算功能，可展示历史自付二和案件自付一负数')
        put_text('2022-02-25：1、加入地铁个人历史自动理算功能（不完善）2、加入地铁个案自动理算功能')
        put_text('2022-02-21：1、优化导出理算结果错误BUG；2、优化地铁赔付查询展示，优化读取速度')
        put_text('2022-02-17：1、修改网络获取时间，批次超1000的案件可以获取不被截断；2、优化筛重检查机制，提升速度；3、导出明细表和理算表恢复名字后面的批次号命名；4、优化错误日志提示，提示更详细')
        put_text('2022-02-15：1、增加地铁全量赔付明细的查询功能')
        put_text('2022-02-13：导入地铁22-01-25之前的全量赔付明细，并提供单人查询功能')
        put_text('2022-02-10：1、增加公交导出明细，同一单位自动合并成一个表格功能；2、单独增加公交社保数据导出理算功能；3、导出理算表时，未理算sheet页增加校验问题件的功能，直接展示是不是问题件')
        put_text('2022-02-09：1、增加‘合计’字段标红的检查；2、优化错误提示显示方式')
        put_text('2022-02-08：修复公交和地铁的单位名称获取BUG；')
    
    列表 = ['登录助手系统1.0', '社保数据TXT文本转成表格', '地铁全量赔付数据', '下载历史文件', '账号换团队', '退单表转换']

    while True:
        选项 = 选择操作项目(列表)

        if 选项 == '登录助手系统1.0':
            账号 = input('请输入你的系统帐号登录😊', type=TEXT, placeholder='是系统的账号哟😀',
                    help_text='首次使用需要注册', required=True)
            # 全局变量
            global headers
            headers = GD.检查登录状态(账号)

            if not os.path.isdir(f'缓存文件夹'):
                os.mkdir(f'缓存文件夹')

            列表 = ['检测基础字段', '检查筛重文件', '批次号导出票据明细', '批次号导出理算结果', '案件号导出票据明细', '案件号导出理算结果', '查询基础信息', '北京人寿自动理算']

            while True:
                选项 = 选择操作项目(列表)

                if 选项 == '检测基础字段':
                    单位 = 按钮选择('请选择单位（公交是按入院时间，地铁是按出院时间）', ['公交', '地铁', '返回'])
                    
                    if 单位 == '返回':
                        continue

                    团批信息 = 是否导入团批()
                    批次号列表 = 批量选择项目('批次号')

                    if 批次号列表:
                        starttime = datetime.now()
                        nmb = 1
                        put_processbar('批次', auto_close=True)
                        for 批次号 in 批次号列表:
                            set_processbar('批次', nmb / len(批次号列表), label=nmb)
                            nmb += 1
                            if 批次号:
                                检查基础字段(批次号, 团批信息, 单位)
                                
                        endtime = datetime.now()
                        put_text(f'本次运行时间为：{endtime - starttime}')

                elif 选项 == '检查筛重文件':
                    try:
                        onefile = 上传单个文件('请上传需要检查的筛重文件，空为返回上一级')
                    except:
                        # popup('没有选择文件，请重新选择。')
                        onefile = ''
                    
                    if onefile:
                        starttime = datetime.now()
                        检查筛重文件(onefile)
                        下载单个文件(onefile)
                                
                        endtime = datetime.now()
                        put_text(f'本次运行时间为：{endtime - starttime}')

                elif 选项 == '批次号导出票据明细':
                    单位 = 按钮选择('请选择单位', ['公交', '地铁', '返回'])
                    
                    if 单位 == '返回':
                        continue

                    批次号列表 = 批量选择项目('批次号')

                    if 批次号列表:
                        starttime = datetime.now()
                        wb = Workbook()     # 创建新工作薄
                        ws = wb.active      # 获取活跃sheet表
                        ws.title = '票据明细'

                        if 单位 == '公交':
                            title = ['姓名', '性别', '身份证号', '案件号', '票据号', '自付一', '起付金额', '超封顶金额', '自付二', '自费', '个人支付', '票据时间', '出院时间', '票据类型', '备注', '单位']

                        else:
                            title = ['姓名', '身份证号', '案件号', '票据号', '自付一', '起付金额', '超封顶金额', '自付二', '自费', '个人支付', '票据时间', '出院时间', '票据类型', '备注', '提交月份']

                        ws.append(title)    # 批量添加标题
                        ws1 = wb.create_sheet(title='问题件')
                        title1 = ['姓名', '身份证号', '案件号', '问题原因', '问题描述', '备注']
                        ws1.append(title1)

                        nmb = 1
                        put_processbar('批次号', auto_close=True)
                        for 批次号 in 批次号列表:
                            set_processbar('批次号', nmb / len(批次号列表), label=nmb)
                            nmb += 1
                            if 批次号:
                                if 单位 == '公交':
                                    # 0是单位简称，1是保单号简称，2是保单号全称
                                    单位信息 = 单位简称获取(批次号)

                                    try:
                                        lst, lstw = 批次号导出票据明细表(批次号, 单位, 单位信息[0])

                                    except Exception as e:
                                        # 输出错误提示
                                        print(datetime.now())
                                        print(traceback.format_exc())
                                        print('====='*50)
                                        print(e)
                                        wb.save(f'缓存文件夹/{单位}票据明细{批次号}.xlsx')
                                        下载单个文件(f'{单位}票据明细{批次号}')
                                        put_text(f'{单位}批次号{批次号}导出错误!!!!!!!!!!!!!!!!!!!!!!!')

                                else:
                                    try:
                                        lst, lstw = 批次号导出票据明细表(批次号, 单位)

                                    except Exception as e:
                                        # 输出错误提示
                                        print(datetime.now())
                                        print(traceback.format_exc())
                                        print('====='*50)
                                        print(e)
                                        wb.save(f'缓存文件夹/{单位}票据明细{批次号}.xlsx')
                                        下载单个文件(f'{单位}票据明细{批次号}')
                                        put_text(f'{单位}批次号{批次号}导出错误!!!!!!!!!!!!!!!!!!!!!!!')
                                
                                for i in lst:
                                    ws.append(i)
                                for j in lstw:
                                    ws1.append(j)

                        wb.save(f'缓存文件夹/{单位}票据明细{批次号}.xlsx')
                        下载单个文件(f'{单位}票据明细{批次号}')
                                
                        endtime = datetime.now()
                        put_text(f'本次运行时间为：{endtime - starttime}')

                elif 选项 == '批次号导出理算结果':
                    单位 = 按钮选择('请选择单位', ['公交', '地铁', '返回'])
                    
                    if 单位 == '返回':
                        continue

                    批次号列表 = 批量选择项目('批次号')

                    if 批次号列表:
                        starttime = datetime.now()
                        wb = Workbook()     # 创建新工作薄
                        ws = wb.active      # 获取活跃sheet表
                        ws.title = '理算结果'
                        if 单位 == '公交':
                            title = ['姓名', '性别', '身份证号', '案件号', '门诊自付一', '门诊起付金额', '门诊超封顶金额', '门诊自付二', '门诊自费', '住院自付一', '住院起付金额', '住院超封顶金额', '住院自付二', '住院自费', '门特自付一', '门特起付金额', '门特超封顶金额', '门特自付二', '门特自费', '门诊理算', '住院理算', '理算结果', '年份', '单位']
                                                    
                        elif 单位 == '地铁':
                            title = ['姓名', '身份证号', '案件号', '门诊自付一', '门诊起付金额', '门诊超封顶金额', '门诊自付二', '门诊自费', '住院自付一', '住院起付金额', '住院超封顶金额', '住院自付二', '住院自费', '门特自付一', '门特起付金额', '门特超封顶金额', '门特自付二', '门特自费', '门诊理算', '住院理算', '理算结果', '年份']

                        ws.append(title)    # 批量添加标题
                        ws1 = wb.create_sheet(title='未理算')
                        title1 = ['姓名', '身份证号', '案件号', '问题原因']
                        ws1.append(title1)

                        nmb = 1
                        put_processbar('批次号', auto_close=True)
                        for 批次号 in 批次号列表:
                            set_processbar('批次号', nmb / len(批次号列表), label=nmb)
                            nmb += 1
                            if 批次号:
                                if 单位 == '公交':
                                    # 0是单位简称，1是保单号简称，2是保单号全称
                                    单位信息 = 单位简称获取(批次号)
                                    try:
                                        lst, lstw = 批次号导出理算结果表(批次号, 单位, 单位信息[0])
                                    except Exception as e:
                                        # 输出错误提示
                                        print(datetime.now())
                                        print(traceback.format_exc())
                                        print('====='*50)
                                        print(e)
                                        wb.save(f'缓存文件夹/{单位}{单位信息[0]}理算{批次号}.xlsx')
                                        下载单个文件(f'{单位}{单位信息[0]}理算{批次号}')
                                        put_text(f'{单位}批次号{批次号}导出错误!!!!!!!!!!!!!!!!!!!!!!!')

                                else:
                                    try:
                                        lst, lstw = 批次号导出理算结果表(批次号, 单位)
                                    except Exception as e:
                                        # 输出错误提示
                                        print(datetime.now())
                                        print(traceback.format_exc())
                                        print('====='*50)
                                        print(e)
                                        wb.save(f'缓存文件夹/{单位}{单位信息[0]}理算明细{批次号}.xlsx')
                                        下载单个文件(f'{单位}{单位信息[0]}理算明细{批次号}')
                                        put_text(f'{单位}批次号{批次号}导出错误!!!!!!!!!!!!!!!!!!!!!!!')

                                for i in lst:
                                    ws.append(i)
                                for j in lstw:
                                    ws1.append(j)
                        
                        wb.save(f'缓存文件夹/{单位}理算明细{批次号}.xlsx')
                        下载单个文件(f'{单位}理算明细{批次号}')
                                
                        endtime = datetime.now()
                        put_text(f'本次运行时间为：{endtime - starttime}')                       

                elif 选项 == '案件号导出票据明细':
                    单位 = 按钮选择('请选择单位', ['公交', '地铁', '返回'])
                    
                    if 单位 == '返回':
                        continue

                    案件号列表 = 批量选择项目('案件号')

                    if 案件号列表:
                        starttime = datetime.now()
                        wb = Workbook()     # 创建新工作薄
                        ws = wb.active      # 获取活跃sheet表
                        ws.title = '票据明细'

                        if 单位 == '公交':
                            title = ['姓名', '性别', '身份证号', '案件号', '票据号', '自付一', '起付金额', '超封顶金额', '自付二', '自费', '个人支付', '票据时间', '出院时间', '票据类型', '备注', '单位']

                        elif 单位 == '地铁':
                            title = ['姓名', '身份证号', '案件号', '票据号', '自付一', '起付金额', '超封顶金额', '自付二', '自费', '个人支付', '票据时间', '出院时间', '票据类型', '备注', '提交月份']

                        ws.append(title)    # 批量添加标题

                        nmb = 1
                        put_processbar('案件号', auto_close=True)
                        for 案件号 in 案件号列表:
                            set_processbar('案件号', nmb / len(案件号列表), label=nmb)
                            nmb += 1
                            if 案件号:
                                if 单位 == '公交':
                                    # 0是单位简称，1是保单号简称，2是保单号全称
                                    单位信息 = 单位简称获取(案件号)

                                    try:
                                        lst = 案件号导出票据明细表(案件号, 单位, 单位信息[0])

                                    except Exception as e:
                                        # 输出错误提示
                                        print(datetime.now())
                                        print(traceback.format_exc())
                                        print('====='*50)
                                        print(e)
                                        wb.save(f'缓存文件夹/{单位}{单位信息[0]}票据-{案件号}.xlsx')
                                        下载单个文件(f'{单位}{单位信息[0]}票据-{案件号}')
                                        put_text(f'{单位}案件号{案件号}导出错误!!!!!!!!!!!!!!!!!!!!!!!')

                                else:
                                    # 0是单位简称，1是保单号简称，2是保单号全称
                                    单位信息 = 单位简称获取(案件号)

                                    try:
                                        lst = 案件号导出票据明细表(案件号, 单位)

                                    except Exception as e:
                                        # 输出错误提示
                                        print(datetime.now())
                                        print(traceback.format_exc())
                                        print('====='*50)
                                        print(e)
                                        wb.save(f'缓存文件夹/{单位}{单位信息[0]}票据-{案件号}.xlsx')
                                        下载单个文件(f'{单位}{单位信息[0]}票据-{案件号}')
                                        put_text(f'{单位}案件号{案件号}导出错误!!!!!!!!!!!!!!!!!!!!!!!')

                                for i in lst:
                                    ws.append(i)

                        wb.save(f'缓存文件夹/{单位}{单位信息[0]}票据-{案件号}.xlsx')
                        下载单个文件(f'{单位}{单位信息[0]}票据-{案件号}')
                                
                        endtime = datetime.now()
                        put_text(f'本次运行时间为：{endtime - starttime}')

                elif 选项 == '案件号导出理算结果':
                    单位 = 按钮选择('请选择单位', ['公交', '地铁', '返回'])
                    
                    if 单位 == '返回':
                        continue

                    案件号列表 = 批量选择项目('案件号')

                    if 案件号列表:
                        starttime = datetime.now()
                        wb = Workbook()     # 创建新工作薄
                        ws = wb.active      # 获取活跃sheet表
                        ws.title = '理算结果'
                        if 单位 == '公交':
                            title = ['姓名', '性别', '身份证号', '案件号', '门诊自付一', '门诊起付金额', '门诊超封顶金额', '门诊自付二', '门诊自费', '住院自付一', '住院起付金额', '住院超封顶金额', '住院自付二', '住院自费', '门特自付一', '门特起付金额', '门特超封顶金额', '门特自付二', '门特自费', '门诊理算', '住院理算', '理算结果', '年份', '单位']
                                                    
                        elif 单位 == '地铁':
                            title = ['姓名', '身份证号', '案件号', '门诊自付一', '门诊起付金额', '门诊超封顶金额', '门诊自付二', '门诊自费', '住院自付一', '住院起付金额', '住院超封顶金额', '住院自付二', '住院自费', '门特自付一', '门特起付金额', '门特超封顶金额', '门特自付二', '门特自费', '门诊理算', '住院理算', '理算结果', '年份']

                        ws.append(title)    # 批量添加标题
                        nmb = 1
                        put_processbar('案件号', auto_close=True)
                        for 案件号 in 案件号列表:
                            set_processbar('案件号', nmb / len(案件号列表), label=nmb)
                            nmb += 1
                            if 案件号:
                                if 单位 == '公交':
                                    # 0是单位简称，1是保单号简称，2是保单号全称
                                    单位信息 = 单位简称获取(案件号)
                                    try:
                                        lst = 案件号导出理算结果表(案件号, 单位, 单位信息[0])
                                    except Exception as e:
                                        # 输出错误提示
                                        print(datetime.now())
                                        print(traceback.format_exc())
                                        print('====='*50)
                                        print(e)
                                        put_text(f'{单位}案件号{案件号}导出错误!!!!!!!!!!!!!!!!!!!!!!!')
                                        wb.save(f'缓存文件夹/{单位}{单位信息[0]}理算-{案件号}.xlsx')
                                        下载单个文件(f'{单位}{单位信息[0]}理算-{案件号}')

                                elif 单位 == '地铁':
                                    # 0是单位简称，1是保单号简称，2是保单号全称
                                    单位信息 = 单位简称获取(案件号)

                                    try:
                                        lst = 案件号导出理算结果表(案件号, 单位)
                                    except Exception as e:
                                        # 输出错误提示
                                        print(datetime.now())
                                        print(traceback.format_exc())
                                        print('====='*50)
                                        print(e)
                                        wb.save(f'缓存文件夹/{单位}{单位信息[0]}理算明细-{案件号}.xlsx')
                                        下载单个文件(f'{单位}{单位信息[0]}理算明细-{案件号}')
                                        put_text(f'{单位}案件号{案件号}导出错误!!!!!!!!!!!!!!!!!!!!!!!')

                                for i in lst:
                                    ws.append(i)

                        wb.save(f'缓存文件夹/{单位}{单位信息[0]}理算明细-{案件号}.xlsx')
                        下载单个文件(f'{单位}{单位信息[0]}理算明细-{案件号}')
                                
                        endtime = datetime.now()
                        put_text(f'本次运行时间为：{endtime - starttime}')

                elif 选项 == '查询基础信息':
                    # while True:
                    # 列表1 = ['批次号导出案件列表信息', '案件号导出上传时间信息', '批次号导出案件数量', '身份证号查询有无身份证影像', '身份证号查询指定上传日期案件号', '返回']
                    选项1 = 按钮选择('选择项目', ['批次号导出案件列表信息', '案件号导出上传时间信息', '批次号导出案件数量', '身份证号查询有无身份证影像', '身份证号查询指定上传日期案件号', '返回'])

                    if 选项1 == '返回':
                        continue

                    elif 选项1 == '批次号导出案件列表信息':
                        选项 = 按钮选择('是否查询时间和理算信息？（选择“是”所有内容导成一个表，选择“否”每个批次一个表）', ['是', '否'])

                        # if 选项 == '是':
                        #     with put_collapse('点击查看导出的内容：'):
                        #         put_table([
                        #             ['序号', '姓名', '身份证号', '案件号', '票据数', '上传时间', '回传时间', '是否理算'],
                        #             ['XXX', 'XXX', 'XXXXXXXX', 'XXXXXX', 'XXXXXX', 'XXXXXXX', 'XXXXXXX', 'XXXXXXX']
                        #         ])
                        # else:
                        #     with put_collapse('点击查看导出的内容：'):
                        #         put_table([
                        #             ['序号', '姓名', '身份证号', '案件号', '票据数'],
                        #             ['XXX', 'XXX', 'XXXXXXXX', 'XXXXXX', 'XXXXXX']
                        #         ])                                

                        批次号列表 = 批量选择项目('批次号')

                        if 批次号列表:
                            starttime = datetime.now()
                            if 选项 == '是':
                                wb = Workbook()     # 创建新工作薄
                                ws = wb.active      # 获取活跃sheet表
                                ws.title = '个案信息明细'
                                title = ['序号', '姓名', '身份证号', '批次号', '案件号', '上传时间', '回传时间', '是否理算']
                                ws.append(title)    # 批量添加标题
                            nmb = 1
                            put_processbar('批次号', auto_close=True)
                            for 批次号 in 批次号列表:
                                set_processbar('批次号', nmb / len(批次号列表), label=nmb)
                                nmb += 1
                                if 批次号:
                                    if 选项 == '否':
                                        wb = Workbook()     # 创建新工作薄
                                        ws = wb.active      # 获取活跃sheet表
                                        ws.title = '个案信息明细'
                                        title = ['序号', '姓名', '身份证号', '案件号', '票据数']
                                        ws.append(title)    # 批量添加标题

                                    try:
                                        lst = 批次号导出案件列表信息(批次号, 选项)
                                        for i in lst:
                                            ws.append(i)

                                        if 选项 == '否':
                                            wb.save(f'缓存文件夹/基础信息{批次号}.xlsx')
                                            下载单个文件(f'基础信息{批次号}')
                            
                                            endtime = datetime.now()
                                            put_text(f'本次运行时间为：{endtime - starttime}')

                                    except Exception as e:
                                        # 输出错误提示
                                        print(datetime.now())
                                        print(traceback.format_exc())
                                        print('====='*50)
                                        print(e)
                                        put_text(f'批次号{批次号}导出错误!!!!!!!!!!!!!!!!!!!!!!!')

                            if 选项 == '是':
                                wb.save(f'缓存文件夹/基础信息{批次号}.xlsx')
                                下载单个文件(f'基础信息{批次号}')
                                        
                                endtime = datetime.now()
                                put_text(f'本次运行时间为：{endtime - starttime}')

                    elif 选项1 == '案件号导出上传时间信息':
                        案件号列表 = 批量选择项目('案件号')

                        if 案件号列表:
                            starttime = datetime.now()
                            wb = Workbook()     # 创建新工作薄
                            ws = wb.active      # 获取活跃sheet表
                            ws.title = '个案信息明细'
                            title = ['序号', '姓名', '身份证号', '批次号', '案件号', '上传时间', '回传时间', '是否理算']
                            ws.append(title)    # 批量添加标题

                            nmb = 1
                            put_processbar('案件号', auto_close=True)
                            for 案件号 in 案件号列表:
                                set_processbar('案件号', nmb / len(案件号列表), label=nmb)
                                if 案件号:
                                    try:
                                        lst = 案件号导出案件列表信息(案件号, nmb)
                                        ws.append(lst)

                                    except Exception as e:
                                        # 输出错误提示
                                        print(datetime.now())
                                        print(traceback.format_exc())
                                        print('====='*50)
                                        print(e)
                                        put_text(f'案件号{案件号}导出错误!!!!!!!!!!!!!!!!!!!!!!!')

                                nmb += 1

                            wb.save(f'缓存文件夹/基础信息{案件号}.xlsx')
                            下载单个文件(f'基础信息{案件号}')
                            
                            endtime = datetime.now()
                            put_text(f'本次运行时间为：{endtime - starttime}')

                    elif 选项1 == '批次号导出案件数量':
                        # with put_collapse('点击查看导出的内容：'):
                            # put_table([
                            #     ['序号', '姓名', '身份证号', '案件号', '票据数'],
                            #     ['XXX', 'XXX', 'XXXXXXXX', 'XXXXXX', '']
                            # ])
                        批次号列表 = 批量选择项目('批次号')

                        if 批次号列表:
                            starttime = datetime.now()
                            wb = Workbook()     # 创建新工作薄
                            ws = wb.active      # 获取活跃sheet表
                            ws.title = '个案信息明细'
                            title = ['序号', '批次号', '案件数量']
                            ws.append(title)    # 批量添加标题
                            nmb = 1
                            put_processbar('批次号', auto_close=True)
                            for 批次号 in 批次号列表:
                                set_processbar('批次号', nmb / len(批次号列表), label=nmb)
                                nmb += 1
                                if 批次号:
                                    try:
                                        url = GD.批次号查询网址(批次号)
                                        data1 = GD.获取案件信息(url, headers)

                                        if data1 == '没有更多啦~':
                                            案件总数 = '-'
                                        else:
                                            案件总数 = data1['page']['count']
                                        lst = [nmb-1, 批次号, 案件总数]
                                        ws.append(lst)

                                    except Exception as e:
                                        # 输出错误提示
                                        print(datetime.now())
                                        print(traceback.format_exc())
                                        print('====='*50)
                                        print(e)
                                        put_text(f'批次号{批次号}导出错误!!!!!!!!!!!!!!!!!!!!!!!')

                            wb.save(f'缓存文件夹/批次号案件数量-{批次号}.xlsx')
                            下载单个文件(f'批次号案件数量-{批次号}')
                            
                            endtime = datetime.now()
                            put_text(f'本次运行时间为：{endtime - starttime}')

                    elif 选项1 == '身份证号查询有无身份证影像':
                        # with put_collapse('点击查看导出的内容：'):
                        #     put_table([
                        #         ['姓名', '身份证号', '有身份证案件号'],
                        #         ['XXX', 'XXXXXXXX', 'XXXXXXXXXXXXX']
                        #     ])
                        身份证号列表 = 批量选择项目('身份证号')
                        if 身份证号列表:
                            starttime = datetime.now()
                            wb = Workbook()     # 创建新工作薄
                            ws = wb.active      # 获取活跃sheet表
                            title = ['姓名', '身份证号', '有身份证案件号']
                            ws.append(title)    # 批量添加标题

                            nmb = 1
                            put_processbar('身份证号', auto_close=True)
                            for 身份证号 in 身份证号列表:
                                set_processbar('身份证号', nmb / len(身份证号列表), label=nmb)
                                nmb += 1
                                if 身份证号:
                                    try:
                                        lst = 个人身份证影像查询(身份证号)
                                        if lst:
                                            for i in lst:
                                                ws.append(i)

                                    except Exception as e:
                                        # 输出错误提示
                                        print(datetime.now())
                                        print(traceback.format_exc())
                                        print('====='*50)
                                        print(e)
                                        wb.save(f'缓存文件夹/身份证影像查询结果-{身份证号}.xlsx')
                                        下载单个文件(f'身份证影像查询结果-{身份证号}')
                                        put_text(f'身份证号{身份证号}导出错误!!!!!!!!!!!!!!!!!!!!!!!')

                            wb.save(f'缓存文件夹/身份证影像查询结果-{身份证号}.xlsx')
                            下载单个文件(f'身份证影像查询结果-{身份证号}')
                            
                            endtime = datetime.now()
                            put_text(f'本次运行时间为：{endtime - starttime}')

                    elif 选项1 == '身份证号查询指定上传日期案件号':
                        # with put_collapse('点击查看导出的内容：'):
                        #     put_table([
                        #         ['姓名', '身份证号', '案件号', '上传日期', '理算状态'],
                        #         ['XXX', 'XXXXXXXX', 'XXXXXX', 'XXXXXXX', 'XXXXXX']
                        #     ])
                        # 选择查询条件
                        查询条件 = input('请输入要查询的上传日期，可以只输入年份或年-月，查询全部案件号直接确认', type=TEXT, placeholder='输入格式：2022 或 2022-02', help_text='注意输入格式')
                        身份证号列表 = 批量选择项目('身份证号')
                        
                        if 身份证号列表:
                            starttime = datetime.now()
                            wb = Workbook()     # 创建新工作薄
                            ws = wb.active      # 获取活跃sheet表
                            ws.append(['姓名', '身份证号', '案件号', '上传日期', '理算状态'])

                            nmb = 1
                            put_processbar('身份证号', auto_close=True)
                            for 身份证号 in 身份证号列表:
                                set_processbar('身份证号', nmb / len(身份证号列表), label=nmb)
                                nmb += 1
                                if 身份证号:
                                    try:
                                        lst = 身份证号指定条件查询案件号(身份证号, 查询条件)
                                        if lst:
                                            for i in lst:
                                                ws.append(i)

                                    except Exception as e:
                                        # 输出错误提示
                                        print(datetime.now())
                                        print(traceback.format_exc())
                                        print('====='*50)
                                        print(e)
                                        wb.save(f'缓存文件夹/案件号查询结果-{身份证号}.xlsx')
                                        下载单个文件(f'案件号查询结果-{身份证号}')
                                        put_text(f'身份证号{身份证号}导出错误!!!!!!!!!!!!!!!!!!!!!!!')

                            wb.save(f'缓存文件夹/案件号查询结果-{身份证号}.xlsx')
                            下载单个文件(f'案件号查询结果-{身份证号}')
                            
                            endtime = datetime.now()
                            put_text(f'本次运行时间为：{endtime - starttime}')

                elif 选项 == '北京人寿自动理算':
                    理算选项 = 按钮选择('请选择项目', ['个人历史案件理算', '单个案件号理算', '批次号理算', '返回'])
                    
                    if 理算选项 == '返回':
                        continue

                    if 理算选项 == '个人历史案件理算':
                        单位 = 按钮选择('请选择单位', ['公交', '地铁', '返回'])
                    
                        if 单位 == '返回':
                            continue

                        查询年份 = 按钮选择('请选择需要查询的年份', ['全部', '20年', '21年', '22年'])
                        获取系统理算 = 按钮选择('是否获取系统的理算', ['是', '否'])
                        if 查询年份 == '全部':
                            查询年份 = ''
                        身份证号列表 = 批量选择项目('身份证号')

                        if 身份证号列表:
                            starttime = datetime.now()
                            wb = Workbook()     # 创建新工作薄
                            ws = wb.active      # 获取活跃sheet表
                            ws.title = '理算结果'


                            if 获取系统理算 == '否':
                                if 单位 == '地铁':
                                    title = ['姓名', '身份证号', '案件号', '上传时间', '回传时间', '门诊理算', '住院理算', '门特理算', '理算合计', '年份', '自付二理算', '方案变更', '本年应赔', '年度已赔']
                                elif 单位 == '公交':
                                    title = ['姓名', '身份证号', '案件号', '上传时间', '回传时间', '门诊理算', '住院理算', '门特理算', '理算合计', '年份']
                                ws.append(title)
                                ws['I1'].fill = PatternFill(patternType="solid", start_color='F5A9BC')  # 自动理算列
                                # ws['J1'].fill = PatternFill(patternType="solid", start_color='FAAC58')  # 系统理算列
                                
                            else:
                                if 单位 == '地铁':
                                    title = ['姓名', '身份证号', '案件号', '上传时间', '回传时间', '门诊理算', '住院理算', '门特理算', '理算合计', '系统门诊', '系统住院', '系统门特', '系统合计', '年份', '自付二理算', '方案变更', '本年应赔', '年度已赔']
                                elif 单位 == '公交':
                                    title = ['姓名', '身份证号', '案件号', '上传时间', '回传时间', '门诊理算', '住院理算', '门特理算', '理算合计', '系统门诊', '系统住院', '系统门特', '系统合计', '年份']
                                ws.append(title)
                                ws['I1'].fill = PatternFill(patternType="solid", start_color='F5A9BC')  # 自动理算列
                                ws['M1'].fill = PatternFill(patternType="solid", start_color='FAAC58')  # 系统理算列


                            nmb = 1
                            put_processbar('身份证号', auto_close=True)
                            for 身份证号 in 身份证号列表:
                                set_processbar('身份证号', nmb / len(身份证号列表), label=nmb)
                                nmb += 1
                                if 身份证号:
                                    try:
                                        lst = 北京人寿个人历史理算查询(身份证号, 单位, 查询年份, 获取系统理算)
                                        if lst:
                                            for i in lst:
                                                ws.append(i)

                                    except Exception as e:
                                        # 输出错误提示
                                        ws.append(['', 身份证号, '理算失败'])
                                        print(datetime.now())
                                        print(traceback.format_exc())
                                        print('='*100)
                                        print(e)
                                        # wb.save(f'缓存文件夹/{单位}自动理算-{身份证号}.xlsx')
                                        # 下载单个文件(f'{单位}自动理算-{身份证号}')
                                        # put_text(f'{单位}身份证号{身份证号}导出错误!!!!!!!!!!!!!!!!!!!!!!!')

                            wb.save(f'缓存文件夹/{单位}自动理算-{身份证号}.xlsx')
                            下载单个文件(f'{单位}自动理算-{身份证号}')
                                
                            endtime = datetime.now()
                            put_text(f'本次运行时间为：{endtime - starttime}')

                    elif 理算选项 == '单个案件号理算':
                        单位 = 按钮选择('请选择单位', ['公交', '地铁', '返回'])
                    
                        if 单位 == '返回':
                            continue

                        案件号列表 = 批量选择项目('案件号')
                        方案编码 = input('请输入编码', required=True)

                        if 案件号列表:
                            starttime = datetime.now()
                            wb = Workbook()     # 创建新工作薄
                            ws = wb.active      # 获取活跃sheet表
                            ws.title = '理算结果'

                            if 单位 == '地铁':
                                title = ['姓名', '身份证号', '案件号', '上传时间', '回传时间', '门诊理算', '住院理算', '门特理算', '理算合计', '系统门诊', '系统住院', '系统门特', '系统合计', '年份']
                            else:
                                title = ['姓名', '身份证号', '案件号', '上传时间', '回传时间', '门诊理算', '住院理算', '门特理算', '理算合计']
                            ws.append(title)    # 批量添加标题

                            nmb = 1
                            put_processbar('案件号', auto_close=True)
                            for 案件号 in 案件号列表:
                                set_processbar('案件号', nmb / len(案件号列表), label=nmb)
                                nmb += 1
                                if 案件号:
                                    if 单位 == '地铁':
                                        try:
                                            lst = 北京人寿地铁个案理算(案件号)
                                            if lst:
                                                for i in lst:
                                                    ws.append(i)

                                        except Exception as e:
                                            # 输出错误提示
                                            print(datetime.now())
                                            print(traceback.format_exc())
                                            print('====='*50)
                                            print(e)
                                            wb.save(f'缓存文件夹/{单位}自动理算-{案件号}.xlsx')
                                            下载单个文件(f'{单位}自动理算-{案件号}')
                                            put_text(f'{单位}案件号{案件号}导出错误!!!!!!!!!!!!!!!!!!!!!!!')

                                    else:
                                        try:
                                            lst = 北京人寿公交问题件理算(案件号, 方案编码)
                                            if lst:
                                                ws.append(lst)

                                        except Exception as e:
                                            # 输出错误提示
                                            print(datetime.now())
                                            print(traceback.format_exc())
                                            print('====='*50)
                                            print(e)
                                            wb.save(f'缓存文件夹/{单位}自动理算-{案件号}.xlsx')
                                            下载单个文件(f'{单位}自动理算-{案件号}')
                                            put_text(f'{单位}案件号{案件号}导出错误!!!!!!!!!!!!!!!!!!!!!!!')
                                        

                            wb.save(f'缓存文件夹/{单位}自动理算-{案件号}.xlsx')
                            下载单个文件(f'{单位}自动理算-{案件号}')
                                
                            endtime = datetime.now()
                            put_text(f'本次运行时间为：{endtime - starttime}')
                            
                    elif 理算选项 == '批次号理算':
                        单位 = 按钮选择('请选择单位', ['公交', '地铁', '返回'])
                    
                        if 单位 == '返回':
                            continue

                        批次号列表 = 批量选择项目('批次号')

                        if 批次号列表:
                            starttime = datetime.now()
                            wb = Workbook()     # 创建新工作薄
                            ws = wb.active      # 获取活跃sheet表
                            ws.title = '历史理算结果'

                            title = ['姓名', '身份证号', '案件号', '上传时间', '回传时间', '门诊理算', '住院理算', '门特理算', '理算合计', '系统门诊', '系统住院', '系统门特', '系统合计', '年份']
                            ws.append(title)    # 批量添加标题

                            nmb = 1
                            put_processbar('批次号', auto_close=True)
                            for 批次号 in 批次号列表:
                                set_processbar('批次号', nmb / len(批次号列表), label=nmb)
                                nmb += 1
                                
                                if 批次号:
                                    try:
                                        lst = 北京人寿批次理算(批次号, 单位)
                                        if lst:
                                            for i in lst:
                                                ws.append(i)

                                    except Exception as e:
                                        # 输出错误提示
                                        print(datetime.now())
                                        print(traceback.format_exc())
                                        print('====='*50)
                                        print(e)
                                        wb.save(f'缓存文件夹/{单位}自动理算-{批次号}.xlsx')
                                        下载单个文件(f'{单位}自动理算-{批次号}')
                                        put_text(f'{单位}案件号{批次号}导出错误!!!!!!!!!!!!!!!!!!!!!!!')

                            wb.save(f'缓存文件夹/{单位}自动理算-{批次号}.xlsx')
                            下载单个文件(f'{单位}自动理算-{批次号}')
                                
                            endtime = datetime.now()
                            put_text(f'本次运行时间为：{endtime - starttime}')
                    
        elif 选项 == '社保数据TXT文本转成表格':
            TXT列表 = 上传多个文件(f'请上传TXT文件')
            保存文件名 = input('输入要保存的文件名')
            
            starttime = datetime.now()
            txts = 文本批量合并(TXT列表)
            if txts:
                txt_xlsx(txts, 保存文件名)
                下载单个文件(保存文件名)
                                
                endtime = datetime.now()
                put_text(f'本次运行时间为：{endtime - starttime}')

        elif 选项 == '下载历史文件':
            选项 = 按钮选择('请选择', ['查看现有历史文件', '查找单个文件'])

            if 选项 == '查看现有历史文件':
                文件列表 = os.listdir(f'缓存文件夹')   #列出指定目录下的所有文件和子目录，包括隐藏文件
                with put_collapse('点击查看历史文件：'):
                    for i in 文件列表:
                        下载单个文件(i)
            elif 选项 == '查找单个文件':
                文件 = input('请输入你要下载的文件名', type=TEXT, required=True)
                文件列表 = os.listdir(f'缓存文件夹')   #列出指定目录下的所有文件和子目录，包括隐藏文件
                for i in 文件列表:
                    if 文件 in i:
                        下载单个文件(i)

        elif 选项 == '账号换团队':
            GD.更改登录团队()

        elif 选项 == '地铁全量赔付数据':
            # 列表 = ['查询数据', '写入数据', '删除数据']
            # 查询数据
            # while True:
            选项2 = 按钮选择('', ['查询数据', '写入数据', '删除数据', '返回'])

            if 选项2 == '返回':
                continue

            elif 选项2 == '查询数据':

                身份证号 = input('请输入身份证号查询个人', type=TEXT, placeholder='只能用身份证号查询')
                地铁全量赔付查询(赔付明细, 身份证号)

            elif 选项2 == '写入数据':
                pass

            elif 选项2 == '删除数据':
                pass

        elif 选项 == '退单表转换':
            列表 = ['地铁超限额退单表', '地铁赔付过万退单表']
            while True:
                选项 = 选择操作项目(列表)

                if 选项 == '地铁超限额退单表':
                    with put_collapse(f'请上传文件，点击查看模板：'):
                        put_table([
                            ['序号', '单位名称', '姓名', '身份证号', '单位已报销金额', '本次已报销', '本次应报销金额', '本年度应报销总额', '在职/退休', '年份', '备注', '案件号'],
                            ['XXXX', 'XXXXXXX', 'XXXX', 'XXXXXXX', 'XXXXXXXXXXXXX', 'XXXXXXXXX', 'XXXXXXXXXXXXX', 'XXXXXXXXXXXXXX', 'XXXXXXXXX', 'XXXX', 'XXX', 'XXXXX'],
                        ])
                    表格路径 = 上传单个文件(f'请上传原始表格')
                    选项 = 按钮选择('是否同时创建个人退单表？', ['是', '否'])

                    if 表格路径:
                        地铁超限额退单表(表格路径, 选项)

                elif 选项 == '地铁赔付过万退单表':
                    with put_collapse(f'请上传文件，点击查看模板：'):
                        put_table([
                            [f'单位名称', '姓名', '身份证号', '案件号'],
                            ['XXXXXXXX', 'XXXX', 'XXXXXXX', 'XXXXX'],
                        ])
                    表格路径 = 上传单个文件(f'请上传原始表格')
                    选项 = 按钮选择('是否同时创建个人退单表？', ['是', '否'])

                    if 表格路径:
                        地铁赔付过万退单表(表格路径, 选项)

if __name__ == '__main__':
    start_server(main, port=8088, debug=False, auto_open_webbrowser=False)