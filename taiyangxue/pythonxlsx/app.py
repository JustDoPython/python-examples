from openpyxl import Workbook

def createWorkbook():
    # 创建一个 workbook 对象
    wb = Workbook()
    # 得到激活的 sheet
    ws = wb.active
    # 单元个设置值
    ws['A1'] = 42
    # 批量设置行
    ws.append([1, 2, 3])

    wb.save('createWorkbook.xlsx')

def loadWorkbook():
    from openpyxl import load_workbook
    wb = load_workbook('sample.xlsx')
    print("sample.xlsx 活动 sheet 第一个单元格值为：", wb.active['A1'].value)

def operateCell():
    wb = Workbook()
    ws = wb.active
    ws.append((1,2,3))
    ws.append((11,22,33))
    ws.append((111,222,333))

    # 操作单列
    for cell in ws["A"]:
        print(cell.value)
    # 操作单行
    for cell in ws["1"]:
        print(cell.value)
    # 操作多列
    for column in ws['A:C']:
        for cell in column:
            print(cell.value)
    # 操作多行
    for row in ws['1:3']:
        for cell in row:
            print(cell.value)
    # 指定范围
    for row in ws['A1:C3']:
        for cell in row:
            print(cell.value)

def setCellFormat():
    # 单元格格式
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
    from openpyxl.styles import numbers

    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value='宋体').font = Font(name=u'宋体', size=12, bold=True, color='FF0000')
    ws.cell(row=2, column=2, value='右对齐').alignment = Alignment(horizontal='right')
    ws.cell(row=3, column=3, value='填充渐变色').fill = PatternFill(fill_type='solid', start_color='FF0000')
    ws.cell(row=4, column=4, value='设置边线').border = Border(left=Side(border_style='thin', color='FF0000'), right= Side(border_style='thin', color='FF0000'))
    ws.cell(row=5, column=5, value='受保护的').protection = Protection(locked=True, hidden=True)
    ws.cell(row=6, column=6, value=0.54).number_format =numbers.FORMAT_PERCENTAGE
    wb.save('setCellFormat.xlsx')
    print('打开文件 setCellFormat.xlsx 查看结果')

def barChart():
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference

    wb = Workbook()
    ws = wb.active

    rows = [
        ('月份', '苹果', '香蕉'),
        (1, 43, 25),
        (2, 10, 30),
        (3, 40, 60),
        (4, 50, 70),
        (5, 20, 10),
        (6, 10, 40),
        (7, 50, 30),
    ]

    for row in rows:
        ws.append(row)

    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "销量柱状图"
    chart1.y_axis.title = '销量'
    chart1.x_axis.title = '月份'

    data = Reference(ws, min_col=2, min_row=1, max_row=8, max_col=3)
    series = Reference(ws, min_col=1, min_row=2, max_row=8)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(series)
    ws.add_chart(chart1, "A10")

    wb.save('barChart.xlsx')
    print('打开文件 barChart.xlsx 查看结果')

def pieChart():
    from openpyxl import Workbook
    from openpyxl.chart import PieChart, Reference

    data = [
        ['水果', '销量'],
        ['苹果', 50],
        ['樱桃', 30],
        ['橘子', 10],
        ['香蕉', 40],
    ]

    wb = Workbook()
    ws = wb.active

    for row in data:
        ws.append(row)

    pie = PieChart()
    labels = Reference(ws, min_col=1, min_row=2, max_row=5)
    data = Reference(ws, min_col=2, min_row=1, max_row=5)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "水果销量占比"

    ws.add_chart(pie, "D1")
    wb.save('piechart.xlsx')
    print('打开文件 piechart.xlsx 查看结果')

if __name__ == '__main__':
    createWorkbook()
    loadWorkbook()
    operateCell()
    setCellFormat()
    barChart()
    pieChart()