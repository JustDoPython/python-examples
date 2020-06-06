#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
@author: 闲欢
"""

import requests
import time
import random
import openpyxl


# 分页获取商品
def get_premium_offer_list(keyword, page):
    offer_list = []
    for i in range(1, int(page) + 1):
        time.sleep(random.randint(0, 10))
        olist = get_page_offer(keyword, i)
        offer_list.extend(olist)
    return offer_list

# 获取一页商品
def get_page_offer(keyword, pageNo):
    url = "https://data.p4psearch.1688.com/data/ajax/get_premium_offer_list.json?beginpage=%d&keywords=%s" % (pageNo, keyword)
    res = requests.get(url)
    result = res.json()
    offerResult = result['data']['content']['offerResult']
    result = []
    for offer in offerResult:
        obj = {}
        # print(offer['attr']['id'])
        obj['id'] = str(offer['attr']['id'])
        # print(offer['title'])
        obj['title'] = str(offer['title']).replace('<font color=red>', '').replace('</font>', '')
        # print(offer['attr']['company']['shopRepurchaseRate'])
        obj['shopRepurchaseRate'] = str(offer['attr']['company']['shopRepurchaseRate'])
        # print(offer['attr']['tradeQuantity']['number'])
        obj['tradeNum'] = int(offer['attr']['tradeQuantity']['number'])
        obj['url'] = str(offer['eurl'])
        result.append(obj)

    return result

# 写Excel
def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        id = value[i].get('id', '')
        title = value[i].get('title', '')
        shopRepurchaseRate = value[i].get('shopRepurchaseRate', '')
        tradeNum = value[i].get('tradeNum', '')
        url = value[i].get('url', '')
        cell = [id, title, shopRepurchaseRate, tradeNum, url]
        sheet.cell(row=1, column=1, value='ID')
        sheet.cell(row=1, column=2, value='标题')
        sheet.cell(row=1, column=3, value='回购率')
        sheet.cell(row=1, column=4, value='成交量')
        sheet.cell(row=1, column=5, value='链接')
        for j in range(0, len(cell)):
            sheet.cell(row=i+2, column=j+1, value=str(cell[j]))
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")


def main(keyword, page):
    offer_list = get_premium_offer_list(keyword, page)
    print(offer_list)
    write_excel_xlsx('./data.xlsx', '数据', offer_list)

if __name__ == '__main__':
    main("数据线", 10)



